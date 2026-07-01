from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from service.route_sheet import route_location_key, upsert_route_location


DEFAULT_WORKBOOK = r"Z:\PEDIDOS SEPID - MGBIO 2025.xlsx"
CONTAINER_Z_ROOT = Path("/mnt/nexora-z")
LOCATIONS_SHEET = "DIRECCIONES"
ROUTE_SHEET = "HOJA DE RUTA"


def _clean(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.upper() == "#N/A" else text


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _resolve_workbook(path_text: str) -> Path:
    path = Path(path_text)
    if path.exists():
        return path
    if len(path_text) >= 2 and path_text[1] == ":" and path_text[0].upper() == "Z":
        relative = path_text[2:].lstrip("\\/")
        mapped = CONTAINER_Z_ROOT / relative.replace("\\", "/")
        if mapped.exists():
            return mapped
    return path


class Command(BaseCommand):
    help = "Importa DIRECCIONES y HOJA DE RUTA desde el Excel operativo"

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Ruta del archivo PEDIDOS SEPID - MGBIO 2025.xlsx")
        parser.add_argument("--from-date", default="2026-06-29", help="Fecha mínima de viaje a importar (YYYY-MM-DD)")
        parser.add_argument("--dry-run", action="store_true", help="Calcula cambios sin escribir en la base")

    def handle(self, *args, **opts):
        workbook_path = _resolve_workbook(str(opts.get("workbook") or DEFAULT_WORKBOOK))
        if not workbook_path.exists():
            raise CommandError(f"No existe el archivo: {workbook_path}")
        from_date = _date(opts.get("from_date"))
        if not from_date:
            raise CommandError("--from-date debe tener formato YYYY-MM-DD")

        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise CommandError(f"No se pudo importar openpyxl: {exc}") from exc

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        if LOCATIONS_SHEET not in wb.sheetnames:
            raise CommandError(f"No existe la hoja {LOCATIONS_SHEET}")
        if ROUTE_SHEET not in wb.sheetnames:
            raise CommandError(f"No existe la hoja {ROUTE_SHEET}")

        locations_ws = wb[LOCATIONS_SHEET]
        route_ws = wb[ROUTE_SHEET]

        location_rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(locations_ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
            name = _clean(row[0])
            if not name:
                continue
            location_rows.append(
                {
                    "name": name,
                    "address": _clean(row[1]),
                    "source_row": row_number,
                }
            )

        stop_rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(route_ws.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
            requester = _clean(row[0])
            requested_date = _date(row[1])
            route_date = _date(row[2])
            time_window = _clean(row[3])
            place_name = _clean(row[4])
            address = _clean(row[5])
            task = _clean(row[6])
            if not route_date or route_date < from_date:
                continue
            if not any([requester, requested_date, time_window, place_name, task]):
                continue
            if not place_name and not task:
                continue
            stop_rows.append(
                {
                    "route_date": route_date,
                    "requested_date": requested_date,
                    "requester_name": requester,
                    "time_window": time_window,
                    "place_name": place_name,
                    "address": address,
                    "task": task,
                    "sort_order": row_number,
                    "source_row": row_number,
                }
            )

        if opts.get("dry_run"):
            self.stdout.write(
                f"DRY RUN: {len(location_rows)} lugares y {len(stop_rows)} paradas desde {from_date.isoformat()}"
            )
            return

        counters = {"locations": 0, "stops_upserted": 0, "stops_skipped_closed": 0}
        with transaction.atomic():
            for item in location_rows:
                upsert_route_location(
                    item["name"],
                    item["address"],
                    source_system="excel",
                    source_sheet=LOCATIONS_SHEET,
                    source_row=item["source_row"],
                    metadata={"workbook": str(workbook_path)},
                )
                counters["locations"] += 1

            with connection.cursor() as cur:
                for item in stop_rows:
                    location_id = None
                    if item["place_name"]:
                        location_id = upsert_route_location(
                            item["place_name"],
                            item["address"],
                            source_system="excel",
                            source_sheet=LOCATIONS_SHEET,
                            source_row=None,
                            metadata={"workbook": str(workbook_path), "routeSheetRow": item["source_row"]},
                        )
                    cur.execute(
                        """
                        INSERT INTO route_stops (
                          id, route_date, requested_date, requester_name, time_window, location_id,
                          place_name, address, task, sort_order, source_system, source_sheet, source_row, metadata
                        )
                        VALUES (
                          %s, %s, %s, %s, %s, %s,
                          %s, %s, %s, %s, 'excel', %s, %s, %s::jsonb
                        )
                        ON CONFLICT (source_system, source_sheet, source_row) DO UPDATE
                        SET route_date = EXCLUDED.route_date,
                            requested_date = EXCLUDED.requested_date,
                            requester_name = EXCLUDED.requester_name,
                            time_window = EXCLUDED.time_window,
                            location_id = EXCLUDED.location_id,
                            place_name = EXCLUDED.place_name,
                            address = EXCLUDED.address,
                            task = EXCLUDED.task,
                            sort_order = EXCLUDED.sort_order,
                            metadata = route_stops.metadata || EXCLUDED.metadata,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE route_stops.status IN ('pendiente','pospuesto')
                        RETURNING id
                        """,
                        [
                            f"rs-{uuid4()}",
                            item["route_date"],
                            item["requested_date"],
                            item["requester_name"],
                            item["time_window"],
                            location_id,
                            item["place_name"],
                            item["address"],
                            item["task"],
                            item["sort_order"],
                            ROUTE_SHEET,
                            item["source_row"],
                            json.dumps(
                                {
                                    "workbook": str(workbook_path),
                                    "locationKey": route_location_key(item["place_name"]),
                                },
                                ensure_ascii=False,
                            ),
                        ],
                    )
                    if cur.fetchone():
                        counters["stops_upserted"] += 1
                    else:
                        counters["stops_skipped_closed"] += 1

        self.stdout.write(
            "IMPORTADO OK: "
            f"{counters['locations']} lugares, "
            f"{counters['stops_upserted']} paradas, "
            f"{counters['stops_skipped_closed']} cerradas sin modificar"
        )
