import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from django.utils import timezone


DEFAULT_WORKBOOK = r"Z:\MG BIO\MADRES\MADRE 2026.xlsx"
INVENTORY_SHEETS = ("SEPID 2026", "POR MG")
MG_RE = re.compile(r"^(MG|NM|NV|CE)\s*0*(\d{1,4})$", re.IGNORECASE)
SERIAL_CLEAN_RE = re.compile(r"[^A-Z0-9]")


def _norm_mg(value):
    raw = "" if value is None else str(value).strip().upper()
    match = MG_RE.match(raw)
    if not match:
        return ""
    return f"{match.group(1).upper()} {int(match.group(2)):04d}"


def _norm_serial(value):
    raw = "" if value is None else str(value).strip().upper()
    return SERIAL_CLEAN_RE.sub("", raw)


def _text(value):
    return "" if value is None else str(value).strip()


def _load_workbook(path):
    if not path.exists():
        raise CommandError(f"No existe el archivo: {path}")
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _load_sales(wb):
    sheet_name = next((name for name in wb.sheetnames if name.strip().upper() == "ALTAS Y BAJAS"), None)
    if not sheet_name:
        raise CommandError("No existe la hoja 'ALTAS Y BAJAS'.")

    sales = {}
    ws = wb[sheet_name]
    for row_num, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        values = list(row[:14])
        code = _norm_mg(values[6] if len(values) > 6 else "")
        motivo = _text(values[10] if len(values) > 10 else "").upper()
        if not code or motivo != "VENTA":
            continue
        serie = _text(values[5] if len(values) > 5 else "")
        sales[code] = {
            "row": row_num,
            "serie": serie,
            "serie_key": _norm_serial(serie),
            "fecha": values[9] if len(values) > 9 else None,
            "observaciones": _text(values[13] if len(values) > 13 else ""),
        }
    return sales


def _load_inventory(wb):
    pairs = {}
    for sheet_name in INVENTORY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_num, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
            values = list(row[:17])
            code = _norm_mg(values[7] if len(values) > 7 else "")
            serie = _text(values[6] if len(values) > 6 else "")
            serie_key = _norm_serial(serie)
            if not code or not serie_key:
                continue
            key = (code, serie_key)
            record = {
                "sheet": sheet_name,
                "row": row_num,
                "prop": _text(values[1] if len(values) > 1 else ""),
                "tipo": _text(values[3] if len(values) > 3 else ""),
                "marca": _text(values[4] if len(values) > 4 else ""),
                "modelo": _text(values[5] if len(values) > 5 else ""),
                "serie": serie,
                "serie_key": serie_key,
                "code": code,
            }
            if key not in pairs or sheet_name == "SEPID 2026":
                pairs[key] = record

    records = list(pairs.values())
    by_code = defaultdict(list)
    by_serial = defaultdict(list)
    for record in records:
        by_code[record["code"]].append(record)
        by_serial[record["serie_key"]].append(record)

    unique_by_serial = {}
    ambiguous_by_serial = {}
    for serie_key, rows in by_serial.items():
        codes = {row["code"] for row in rows}
        if len(codes) == 1:
            unique_by_serial[serie_key] = rows[0]
        else:
            ambiguous_by_serial[serie_key] = rows

    return {
        "records": records,
        "by_code": dict(by_code),
        "unique_by_serial": unique_by_serial,
        "ambiguous_by_serial": ambiguous_by_serial,
    }


def _find_customer_id(cur, hint):
    value = _text(hint)
    if not value:
        return None
    aliases = {
        "PUGLISI": "%PUGLISI%",
        "NOVAMED": "%NOVAMED%",
        "TMD": "%TMD%",
        "RESPILIFE": "%RESPILIFE%",
        "IBRAHIM": "%BRAHIM%",
    }
    upper = value.upper()
    pattern = None
    for key, candidate in aliases.items():
        if key in upper:
            pattern = candidate
            break
    if pattern is None:
        pattern = f"%{value}%"
    cur.execute(
        "SELECT id FROM customers WHERE razon_social ILIKE %s ORDER BY id LIMIT 1",
        [pattern],
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def _fetch_device_rows(cur):
    cur.execute(
        """
        SELECT
          d.id,
          COALESCE(d.numero_interno, '') AS numero_interno,
          COALESCE(d.numero_serie, '') AS numero_serie,
          COALESCE(d.n_de_control, '') AS n_de_control,
          COALESCE(d.mg_estado, 'activo') AS mg_estado,
          COALESCE(c.razon_social, '') AS cliente
        FROM devices d
        LEFT JOIN customers c ON c.id = d.customer_id
        ORDER BY d.id
        """
    )
    rows = []
    for row in cur.fetchall():
        numero_interno = row[1]
        numero_serie = row[2]
        n_de_control = row[3]
        rows.append(
            {
                "id": int(row[0]),
                "numero_interno": numero_interno,
                "numero_serie": numero_serie,
                "n_de_control": n_de_control,
                "mg_estado": row[4],
                "cliente": row[5],
                "numero_interno_code": _norm_mg(numero_interno),
                "n_de_control_code": _norm_mg(n_de_control),
                "serie_key": _norm_serial(numero_serie),
            }
        )
    return rows


def _dash_location_id(cur):
    cur.execute("SELECT id FROM locations WHERE nombre='-' ORDER BY id LIMIT 1")
    row = cur.fetchone()
    return int(row[0]) if row else None


def _effective_code(row, inventory_updates_by_id):
    inventory_record = inventory_updates_by_id.get(row["id"])
    if inventory_record:
        return inventory_record["code"]
    return row["numero_interno_code"] or row["n_de_control_code"]


def _prepare_inventory_updates(device_rows, inventory):
    updates = []
    skipped_ambiguous = []
    unique_by_serial = inventory["unique_by_serial"]
    ambiguous_by_serial = inventory["ambiguous_by_serial"]

    for row in device_rows:
        serie_key = row["serie_key"]
        if not serie_key:
            continue
        if serie_key in ambiguous_by_serial:
            codes = {record["code"] for record in ambiguous_by_serial[serie_key]}
            if row["numero_interno_code"] not in codes:
                skipped_ambiguous.append((row, ambiguous_by_serial[serie_key]))
            continue
        inventory_record = unique_by_serial.get(serie_key)
        if not inventory_record:
            continue
        if row["numero_interno_code"] != inventory_record["code"]:
            updates.append((row, inventory_record))

    return updates, skipped_ambiguous


def _split_inventory_conflicts(inventory_updates, device_rows):
    by_numero_interno = defaultdict(list)
    for row in device_rows:
        if row["numero_interno_code"]:
            by_numero_interno[row["numero_interno_code"]].append(row)

    safe_updates = []
    conflicts = []
    for row, inventory_record in inventory_updates:
        blockers = [
            other for other in by_numero_interno.get(inventory_record["code"], [])
            if other["id"] != row["id"]
        ]
        if blockers:
            conflicts.append((row, inventory_record, blockers))
        else:
            safe_updates.append((row, inventory_record))
    return safe_updates, conflicts


class Command(BaseCommand):
    help = "Sincroniza inventario, propiedad y ventas MG usando MADRE 2026.xlsx como fuente de control."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Ruta al archivo MADRE 2026.xlsx.")
        parser.add_argument("--apply", action="store_true", help="Aplica las correcciones detectadas.")

    def handle(self, *args, **opts):
        workbook_path = Path(opts["workbook"])
        apply_changes = bool(opts.get("apply"))
        wb = _load_workbook(workbook_path)
        sales = _load_sales(wb)
        inventory = _load_inventory(wb)

        with connection.cursor() as cur:
            device_rows = _fetch_device_rows(cur)
            inventory_updates, skipped_ambiguous = _prepare_inventory_updates(device_rows, inventory)
            inventory_updates, inventory_conflicts = _split_inventory_conflicts(inventory_updates, device_rows)
            inventory_updates_by_id = {
                row["id"]: inventory_record for row, inventory_record in inventory_updates
            }

            db_by_code = defaultdict(list)
            for row in device_rows:
                code = _effective_code(row, inventory_updates_by_id)
                if code:
                    row = {**row, "effective_code": code}
                    db_by_code[code].append(row)

            mark_sold = []
            reactivate = []
            missing = []

            for code, sale in sales.items():
                rows = db_by_code.get(code) or []
                exact = [
                    row for row in rows
                    if row["serie_key"] and row["serie_key"] == sale["serie_key"]
                ]
                if not exact:
                    missing.append((code, sale))
                    continue
                for row in exact:
                    if row["mg_estado"] != "inactivo_venta":
                        buyer_id = _find_customer_id(cur, sale["observaciones"])
                        mark_sold.append((code, sale, row, buyer_id))

            for row in device_rows:
                code = _effective_code(row, inventory_updates_by_id)
                if not code or code in sales or row["mg_estado"] != "inactivo_venta":
                    continue
                inventory_records = inventory["by_code"].get(code) or []
                if not any(record["serie_key"] == row["serie_key"] for record in inventory_records):
                    continue
                reactivate.append((code, row, inventory_updates_by_id.get(row["id"]) or inventory_records[0]))

            self.stdout.write(
                "Inventario vigente Excel: "
                f"{len(inventory['records'])} pares stock/serie, "
                f"{len(inventory['unique_by_serial'])} series únicas, "
                f"{len(inventory['ambiguous_by_serial'])} series ambiguas."
            )

            self.stdout.write(f"Para cargar/corregir número interno desde MADRE: {len(inventory_updates)}")
            for row, inventory_record in inventory_updates:
                before = row["numero_interno"] or "-"
                self.stdout.write(
                    f"- device #{row['id']} serie {row['numero_serie']} "
                    f"{before} -> {inventory_record['code']} "
                    f"({inventory_record['sheet']} fila {inventory_record['row']}) "
                    f"cliente '{row['cliente'] or '-'}'"
                )

            self.stdout.write(f"Series ambiguas omitidas: {len(skipped_ambiguous)}")
            for row, records in skipped_ambiguous:
                codes = ", ".join(sorted({record["code"] for record in records}))
                self.stdout.write(
                    f"- device #{row['id']} serie {row['numero_serie']} "
                    f"coincide con más de un stock: {codes}"
                )

            self.stdout.write(f"Conflictos de número interno omitidos: {len(inventory_conflicts)}")
            for row, inventory_record, blockers in inventory_conflicts:
                blocker_text = ", ".join(
                    f"#{other['id']} serie {other['numero_serie'] or '-'} cliente '{other['cliente'] or '-'}'"
                    for other in blockers
                )
                self.stdout.write(
                    f"- device #{row['id']} serie {row['numero_serie']} -> {inventory_record['code']} "
                    f"bloqueado por {blocker_text}"
                )

            self.stdout.write(f"Ventas en Excel: {len(sales)}")
            self.stdout.write(f"Para marcar vendidos: {len(mark_sold)}")
            for code, sale, row, buyer_id in mark_sold:
                self.stdout.write(
                    f"- {code} device #{row['id']} serie {row['numero_serie']} "
                    f"fila {sale['row']} comprador '{sale['observaciones'] or '-'}' "
                    f"customer_id {buyer_id or '-'}"
                )

            self.stdout.write(f"Para reactivar por falso vendido: {len(reactivate)}")
            for code, row, inv in reactivate:
                self.stdout.write(
                    f"- {code} device #{row['id']} serie {row['numero_serie']} "
                    f"figura en {inv['sheet']} fila {inv['row']}"
                )

            self.stdout.write(f"Ventas Excel sin device encontrado por stock+serie: {len(missing)}")
            for code, sale in missing:
                self.stdout.write(f"- {code} fila {sale['row']} serie {sale['serie']}")

            if not apply_changes:
                self.stdout.write("Dry-run finalizado. Ejecute con --apply para aplicar.")
                return

            now_ts = timezone.now()
            dash_id = _dash_location_id(cur)
            with transaction.atomic():
                for row, inventory_record in inventory_updates:
                    cur.execute(
                        "UPDATE devices SET numero_interno = %s WHERE id = %s",
                        [inventory_record["code"], row["id"]],
                    )

                for code, sale, row, buyer_id in mark_sold:
                    sale_date = sale["fecha"] or now_ts
                    obs = f"MADRE 2026 ALTAS Y BAJAS fila {sale['row']}: VENTA"
                    if sale["observaciones"]:
                        obs = f"{obs} - {sale['observaciones']}"
                    sets = [
                        "numero_interno = %s",
                        "mg_estado = 'inactivo_venta'",
                        "mg_inactivo_desde = %s",
                        "mg_venta_fecha = %s",
                        "mg_venta_factura_numero = NULL",
                        "mg_venta_remito_numero = NULL",
                        "mg_venta_observaciones = %s",
                        "mg_venta_usuario_id = NULL",
                        "mg_venta_customer_id = %s",
                        "mg_venta_numero_alternativo = NULL",
                        "alquilado = FALSE",
                    ]
                    params = [code, sale_date, sale_date, obs, buyer_id]
                    if dash_id is not None:
                        sets.append("ubicacion_id = %s")
                        params.append(dash_id)
                    params.append(row["id"])
                    cur.execute(f"UPDATE devices SET {', '.join(sets)} WHERE id = %s", params)
                    cur.execute(
                        """
                        INSERT INTO device_mg_events(
                          device_id, accion, numero_interno_snapshot, fecha_evento,
                          factura_numero, remito_numero, observaciones, usuario_id, ingreso_id,
                          venta_customer_id, venta_numero_alternativo, source
                        )
                        VALUES (%s, 'venta', %s, %s, NULL, NULL, %s, NULL, NULL, %s, NULL, 'equipos')
                        """,
                        [row["id"], code, sale_date, obs, buyer_id],
                    )

                for code, row, inv in reactivate:
                    obs = (
                        f"Corrección MADRE 2026: figura vigente en {inv['sheet']} "
                        f"fila {inv['row']} y no está en ALTAS Y BAJAS como VENTA."
                    )
                    cur.execute(
                        """
                        UPDATE devices
                           SET numero_interno = %s,
                               mg_estado = 'activo',
                               mg_inactivo_desde = NULL,
                               mg_venta_fecha = NULL,
                               mg_venta_factura_numero = NULL,
                               mg_venta_remito_numero = NULL,
                               mg_venta_observaciones = NULL,
                               mg_venta_usuario_id = NULL,
                               mg_venta_customer_id = NULL,
                               mg_venta_numero_alternativo = NULL
                         WHERE id = %s
                        """,
                        [code, row["id"]],
                    )
                    cur.execute(
                        """
                        INSERT INTO device_mg_events(
                          device_id, accion, numero_interno_snapshot, fecha_evento,
                          factura_numero, remito_numero, observaciones, usuario_id, ingreso_id,
                          venta_customer_id, venta_numero_alternativo, source
                        )
                        VALUES (%s, 'reactivacion', %s, %s, NULL, NULL, %s, NULL, NULL, NULL, NULL, 'equipos')
                        """,
                        [row["id"], code, now_ts, obs],
                    )

        self.stdout.write(
            "APLICADO OK: "
            f"números internos corregidos {len(inventory_updates)}, "
            f"vendidos marcados {len(mark_sold)}, "
            f"reactivados {len(reactivate)}, "
            f"ventas faltantes {len(missing)}."
        )
