from io import BytesIO
from unittest.mock import MagicMock, patch

from django.test import SimpleTestCase
from openpyxl import load_workbook
from rest_framework.test import APIRequestFactory, force_authenticate

from service.views.ingresos_views import PresupuestadosExportView


class PresupuestadosExportViewTests(SimpleTestCase):
    def setUp(self):
        self.user = type("AuthUser", (), {"is_authenticated": True, "id": 1})()
        self.factory = APIRequestFactory()

    def test_export_omite_fecha_emision_y_formatea_monto_sin_iva(self):
        request = self.factory.get("/api/ingresos/presupuestados/export/?ids=29321")
        force_authenticate(request, user=self.user)

        cursor = MagicMock()
        cursor_context = MagicMock()
        cursor_context.__enter__.return_value = cursor
        cursor_context.__exit__.return_value = None
        connection = MagicMock()
        connection.cursor.return_value = cursor_context
        rows = [
            {
                "id": 29321,
                "cliente": "TMD",
                "tipo_equipo": "CALENTADOR HUMIDIFICADOR",
                "marca": "Marbel",
                "modelo": "C-5",
                "equipo_variante": "",
                "numero_serie": "D3-1553",
                "numero_interno": "",
                "subtotal_sin_iva": 25169.25,
                "mg_estado": "activo",
                "mg_inactivo_venta": False,
            }
        ]

        with (
            patch("service.views.ingresos_views._mg_list_select_sql", return_value=""),
            patch("service.views.ingresos_views._set_audit_user"),
            patch("service.views.ingresos_views.connection", connection),
            patch("service.views.ingresos_views._fetchall_dicts", return_value=rows),
        ):
            response = PresupuestadosExportView.as_view()(request)

        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook.active

        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["OS", "Cliente", "Equipo", "N/S", "Monto sin IVA"],
        )
        self.assertEqual(sheet.max_column, 5)
        self.assertAlmostEqual(sheet["E2"].value, 25169.25)
        self.assertEqual(sheet["E2"].number_format, '"$" #,##0.00')
