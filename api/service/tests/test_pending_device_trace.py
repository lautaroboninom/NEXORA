import csv
import json
import tempfile
from datetime import datetime
from pathlib import Path
from unittest import skipUnless
from unittest.mock import patch

from django.core.management import call_command
from django.db import connection
from django.test import SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import Workbook

from service.models import User
from service.pending_device_trace import (
    WorkbookRecord,
    _build_access_evidence,
    _index_access_fallback_rows,
    _select_access_row,
    build_proposal,
    load_novamed_workbook,
    serial_search_variants,
)


class PendingDeviceTraceHelpersTest(SimpleTestCase):
    def _candidate(self, **extra):
        base = {
            "ingreso_id": 10,
            "device_id": 20,
            "numero_serie": "N5143421",
            "numero_interno": "MG 1001",
            "cliente": "Cliente Demo",
            "marca": "AirSep",
            "modelo": "New Life",
            "tipo_equipo": "Concentrador",
            "variante": "",
            "mg_estado": "activo",
        }
        base.update(extra)
        return base

    def test_serial_search_variants_airsep_supports_with_and_without_n(self):
        variants = serial_search_variants(self._candidate())
        self.assertEqual([item["value"] for item in variants], ["N5143421", "5143421"])
        self.assertEqual(variants[1]["match_type"], "serial_variant_sin_n")

    def test_build_proposal_access_only_is_revision_manual_low(self):
        proposal = build_proposal(
            candidate=self._candidate(numero_serie="ABC123"),
            workbook_matches={"sales": [], "rentals": [], "loans": []},
            internal_summary={
                "sale_event": False,
                "sale_event_date": None,
                "sale_event_factura": "",
                "sale_event_remito": "",
                "baja_historical": False,
                "baja_historical_date": None,
                "alquiler_historical": False,
                "alquiler_historical_date": None,
                "alquiler_historical_payload": {},
                "terminal_event": False,
                "terminal_event_state": "",
                "terminal_event_date": None,
            },
            access_row={"Id": "10", "NumeroSerie": "ABC123", "Entregado": "True"},
            bejerman_evidences=[],
            duplicate_serials={},
            sale_customer_id=None,
        )
        self.assertEqual(proposal["action"], "revision_manual")
        self.assertEqual(proposal["confidence"], "baja")
        self.assertIn("Access", proposal["reason"])

    def test_build_proposal_access_baja_suggests_baja(self):
        proposal = build_proposal(
            candidate=self._candidate(numero_serie="ABC123", numero_interno="MG 0966"),
            workbook_matches={"sales": [], "rentals": [], "loans": []},
            internal_summary={
                "sale_event": False,
                "sale_event_date": None,
                "sale_event_factura": "",
                "sale_event_remito": "",
                "baja_historical": False,
                "baja_historical_date": None,
                "alquiler_historical": False,
                "alquiler_historical_date": None,
                "alquiler_historical_payload": {},
                "terminal_event": False,
                "terminal_event_state": "",
                "terminal_event_date": None,
            },
            access_row={
                "Id": "10",
                "NumeroSerie": "MG 0966",
                "Recibido": "BAJA",
                "Comentarios": "",
                "FechaIngreso": "2023-08-08 00:00:00",
            },
            bejerman_evidences=[],
            duplicate_serials={},
            sale_customer_id=None,
        )
        self.assertEqual(proposal["action"], "baja")
        self.assertEqual(proposal["confidence"], "baja")
        self.assertIn("BAJA", proposal["reason"])

    def test_build_proposal_rental_without_date_is_revision_manual(self):
        rental_record = WorkbookRecord(
            sheet_key="alquileres_sepid",
            source_name="alquileres_sepid",
            row_number=3,
            codigo_gen="1101004",
            fecha=None,
            comprobante="",
            equipo="Aspirador",
            serial_raw="ABC123",
            serial_norm="ABC123",
            precio_neto="50",
            iva="0.21",
            remito="",
        )
        proposal = build_proposal(
            candidate=self._candidate(numero_serie="ABC123", marca="Silfab", modelo="N-33"),
            workbook_matches={
                "sales": [],
                "rentals": [{"record": rental_record, "serial_info": {"match_type": "serial_exact"}}],
                "loans": [],
            },
            internal_summary={
                "sale_event": False,
                "sale_event_date": None,
                "sale_event_factura": "",
                "sale_event_remito": "",
                "baja_historical": False,
                "baja_historical_date": None,
                "alquiler_historical": False,
                "alquiler_historical_date": None,
                "alquiler_historical_payload": {},
                "terminal_event": False,
                "terminal_event_state": "",
                "terminal_event_date": None,
            },
            access_row=None,
            bejerman_evidences=[],
            duplicate_serials={},
            sale_customer_id=None,
        )
        self.assertEqual(proposal["action"], "revision_manual")
        self.assertEqual(proposal["confidence"], "media")
        self.assertIn("sin fecha efectiva", proposal["reason"].lower())

    def test_build_proposal_prestamo_becomes_alquiler(self):
        loan_record = WorkbookRecord(
            sheet_key="prestamo",
            source_name="prestamo",
            row_number=5,
            codigo_gen="1115007",
            fecha=None,
            comprobante="",
            equipo="Concentrador",
            serial_raw="N5167674",
            serial_norm="N5167674",
            precio_neto="-",
            iva="RT 25696 SEPID",
            remito="",
        )
        proposal = build_proposal(
            candidate=self._candidate(numero_serie="N5167674", numero_interno="MG 2723"),
            workbook_matches={
                "sales": [],
                "rentals": [],
                "loans": [{"record": loan_record, "serial_info": {"match_type": "serial_exact"}}],
            },
            internal_summary={
                "sale_event": False,
                "sale_event_date": None,
                "sale_event_factura": "",
                "sale_event_remito": "",
                "baja_historical": False,
                "baja_historical_date": None,
                "alquiler_historical": False,
                "alquiler_historical_date": None,
                "alquiler_historical_payload": {},
                "terminal_event": False,
                "terminal_event_state": "",
                "terminal_event_date": None,
            },
            access_row=None,
            bejerman_evidences=[],
            duplicate_serials={},
            sale_customer_id=None,
        )
        self.assertEqual(proposal["action"], "alquilado")
        self.assertEqual(proposal["confidence"], "media")
        self.assertEqual(proposal["alquiler_a"], "NOVAMED SA")
        self.assertIn("falta fecha efectiva", proposal["reason"].lower())

    def test_build_access_evidence_accepts_mg_in_access_numero_serie(self):
        evidences, conflicts = _build_access_evidence(
            self._candidate(ingreso_id=22158, device_id=18470, numero_serie="06-0142-09A", numero_interno="MG 0966"),
            {
                "Id": "22158",
                "NumeroSerie": "MG 0966",
                "NdeControl": "",
                "Entregado": "True",
                "Alquilado": "False",
                "Venta": "False",
                "Estado": "8",
                "FechaEntrega": "2023-08-08 00:00:00",
                "FechaIngreso": "2020-10-26 00:00:00",
                "Factura": "",
                "Remito": "0",
                "Recibido": "BAJA",
                "Comentarios": "",
                "RecibeAlquiler": "",
                "CargoAlquiler": "",
            },
        )
        self.assertEqual(conflicts, [])
        self.assertEqual(len(evidences), 1)
        self.assertEqual(evidences[0]["match_type"], "os_exact_mg_en_numero_serie")
        self.assertIn("usa el MG", evidences[0]["detalle"])

    def test_select_access_row_accepts_remito_shifted_when_mg_matches(self):
        candidate = self._candidate(
            ingreso_id=28552,
            device_id=28602,
            numero_serie="",
            numero_interno="",
            n_de_control="MG 3216",
            marca="Meditech",
            modelo="G3G",
            remito_ingreso="25604",
        )
        fallback_index = _index_access_fallback_rows(
            [
                {
                    "Id": "28553",
                    "NumeroSerie": "MG 3216",
                    "NdeControl": "",
                    "RemitoIngreso": "25604",
                    "Remito": "",
                    "Marca": "MEDITECH",
                    "Modelo": "G3G",
                    "Entregado": "False",
                    "Alquilado": "False",
                    "Venta": "False",
                    "FechaIngreso": "2025-09-25 00:00:00",
                    "Recibido": "",
                    "Comentarios": "",
                    "RecibeAlquiler": "",
                    "CargoAlquiler": "",
                }
            ]
        )
        access_row = _select_access_row(candidate, None, fallback_index)
        self.assertIsNotNone(access_row)
        evidences, conflicts = _build_access_evidence(candidate, access_row)
        self.assertEqual(conflicts, [])
        self.assertEqual(evidences[0]["match_type"], "remito_ingreso_mg_shifted")
        self.assertIn("remito de ingreso 25604", evidences[0]["detalle"])

    def test_load_novamed_workbook_reads_expected_rows(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "novamed.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "VENTAS NOVAMED - MG BIO SA"
            ws.append(["Cód. Gen.", "Fecha FC", "Comp.", "Equipo", "N/S", "Precio Neto", "IVA", "Comp. - Nro. Remito"])
            ws.append(["1115007", "09/04/2025", "FC A 0001", "Concentrador", "N5143421", 1, 0.21, "0001"])
            ws2 = workbook.create_sheet("VENTAS NOVAMED - SEPID SA")
            ws2.append(["Cód. Gen.", "Fecha FC", "Comp.", "Equipo", "N/S", "Precio Neto", "IVA", "Comp. - Nro. Remito"])
            ws2.append(["1115003", "-", "PENDIENTE", "Longfian", "MZJ5S121687", 1, 0.21, "0002"])
            ws3 = workbook.create_sheet("ALQUILERES NOVAMED - SEPID SA")
            ws3.append(["Cód. Gen.", "Equipo", "N/S", "Precio Neto", "IVA"])
            ws3.append(["1101004", "Aspirador", "ABC123", 50, 0.21])
            ws4 = workbook.create_sheet("PRÉSTAMO")
            ws4.append(["Cód. Gen.", "Equipo", "N/S", "ID"])
            ws4.append(["1115007", "Concentrador", "N5137572", "-"])
            workbook.save(path)

            index = load_novamed_workbook(str(path))

        self.assertEqual(len(index.lookup("ventas_mg", "N5143421")), 1)
        self.assertEqual(len(index.lookup("ventas_sepid", "MZJ5S121687")), 1)
        self.assertEqual(len(index.lookup("alquileres_sepid", "ABC123")), 1)
        self.assertEqual(len(index.lookup("prestamo", "N5137572")), 1)


@skipUnless(connection.vendor == "postgresql", "Requiere PostgreSQL")
class PendingDeviceTraceCommandsTest(TestCase):
    @classmethod
    def setUpClass(cls):
        with connection.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id BIGSERIAL PRIMARY KEY,
                    nombre TEXT,
                    email VARCHAR(320) UNIQUE,
                    hash_pw TEXT,
                    rol TEXT,
                    activo BOOLEAN DEFAULT TRUE
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS customers (
                    id BIGSERIAL PRIMARY KEY,
                    cod_empresa TEXT,
                    razon_social TEXT NOT NULL,
                    telefono TEXT
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS marcas (
                    id BIGSERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS models (
                    id BIGSERIAL PRIMARY KEY,
                    marca_id INTEGER REFERENCES marcas(id),
                    nombre TEXT NOT NULL,
                    tipo_equipo TEXT,
                    variante TEXT
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS locations (
                    id BIGSERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS devices (
                    id BIGSERIAL PRIMARY KEY,
                    customer_id INTEGER NOT NULL REFERENCES customers(id),
                    marca_id INTEGER REFERENCES marcas(id),
                    model_id INTEGER REFERENCES models(id),
                    numero_serie TEXT,
                    numero_interno TEXT,
                    n_de_control TEXT,
                    tipo_equipo TEXT,
                    variante TEXT,
                    alquilado BOOLEAN NOT NULL DEFAULT FALSE,
                    alquiler_a TEXT,
                    ubicacion_id INTEGER NULL REFERENCES locations(id)
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS ingresos (
                    id BIGSERIAL PRIMARY KEY,
                    device_id INTEGER NOT NULL REFERENCES devices(id),
                    estado TEXT,
                    motivo TEXT,
                    fecha_ingreso TIMESTAMPTZ NULL,
                    fecha_creacion TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    fecha_entrega TIMESTAMPTZ NULL,
                    ubicacion_id INTEGER NULL REFERENCES locations(id),
                    remito_salida TEXT,
                    factura_numero TEXT,
                    alquilado BOOLEAN NOT NULL DEFAULT FALSE,
                    alquiler_a TEXT,
                    alquiler_remito TEXT,
                    alquiler_fecha DATE
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS ingreso_events (
                    id BIGSERIAL PRIMARY KEY,
                    ticket_id INTEGER,
                    ingreso_id INTEGER,
                    de_estado TEXT,
                    a_estado TEXT,
                    usuario_id INTEGER,
                    ts TIMESTAMPTZ,
                    comentario TEXT
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS device_mg_events (
                    id BIGSERIAL PRIMARY KEY,
                    device_id INTEGER NOT NULL REFERENCES devices(id) ON DELETE CASCADE,
                    accion TEXT NOT NULL,
                    numero_interno_snapshot TEXT,
                    fecha_evento TIMESTAMPTZ,
                    factura_numero TEXT,
                    remito_numero TEXT,
                    observaciones TEXT,
                    usuario_id INTEGER,
                    ingreso_id INTEGER,
                    source TEXT DEFAULT 'equipos'
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS audit_log (
                    id BIGSERIAL PRIMARY KEY,
                    ts TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    user_id INTEGER NULL,
                    role TEXT,
                    method TEXT NOT NULL,
                    path TEXT NOT NULL,
                    ip TEXT,
                    user_agent TEXT,
                    status_code INTEGER NOT NULL,
                    body JSONB
                )
                """
            )
            cur.execute("CREATE SCHEMA IF NOT EXISTS audit")
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS audit.change_log (
                    id BIGSERIAL PRIMARY KEY,
                    ts TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    user_id INTEGER NULL,
                    user_role TEXT,
                    table_name TEXT,
                    record_id INTEGER,
                    column_name TEXT,
                    old_value TEXT,
                    new_value TEXT,
                    ingreso_id INTEGER
                )
                """
            )
        call_command("apply_mg_sale_schema")
        call_command("apply_historical_corrections_schema")
        super().setUpClass()

    @classmethod
    def _last_id(cls, cur):
        cur.execute("SELECT LASTVAL()")
        return int(cur.fetchone()[0])

    @classmethod
    def setUpTestData(cls):
        with connection.cursor() as cur:
            cur.execute("DELETE FROM ingreso_historical_corrections")
            cur.execute("DELETE FROM device_mg_events")
            cur.execute("DELETE FROM ingreso_events")
            cur.execute("DELETE FROM audit.change_log")
            cur.execute("DELETE FROM audit_log")
            cur.execute("DELETE FROM ingresos")
            cur.execute("DELETE FROM devices")
            cur.execute("DELETE FROM models")
            cur.execute("DELETE FROM marcas")
            cur.execute("DELETE FROM locations")
            cur.execute("DELETE FROM customers")
        User.objects.all().delete()

        cls.actor = User.objects.create(
            nombre="Admin Traza",
            email="admin-traza@example.com",
            hash_pw="",
            rol="admin",
            activo=True,
        )
        with connection.cursor() as cur:
            cur.execute("INSERT INTO customers(cod_empresa, razon_social, telefono) VALUES (%s,%s,%s)", ["MGBIO", "MG BIO", ""])
            cls.customer_mg = cls._last_id(cur)
            cur.execute("INSERT INTO customers(cod_empresa, razon_social, telefono) VALUES (%s,%s,%s)", ["NOVA", "NOVAMED SA", ""])
            cls.customer_novamed = cls._last_id(cur)
            cur.execute("INSERT INTO customers(cod_empresa, razon_social, telefono) VALUES (%s,%s,%s)", ["CLI1", "Cliente Demo", ""])
            cls.customer_cli = cls._last_id(cur)
            cur.execute("INSERT INTO marcas(nombre) VALUES (%s)", ["AirSep"])
            cls.marca_id = cls._last_id(cur)
            cur.execute(
                "INSERT INTO models(marca_id, nombre, tipo_equipo, variante) VALUES (%s,%s,%s,%s)",
                [cls.marca_id, "New Life", "Concentrador", ""],
            )
            cls.model_id = cls._last_id(cur)
            cur.execute("INSERT INTO locations(nombre) VALUES (%s)", ["-"])
            cls.loc_dash = cls._last_id(cur)
            cur.execute("INSERT INTO locations(nombre) VALUES (%s)", ["Taller"])
            cls.loc_taller = cls._last_id(cur)

    def setUp(self):
        super().setUp()
        with connection.cursor() as cur:
            cur.execute("DELETE FROM ingreso_historical_corrections")
            cur.execute("DELETE FROM device_mg_events")
            cur.execute("DELETE FROM ingreso_events")
            cur.execute("DELETE FROM audit.change_log")
            cur.execute("DELETE FROM audit_log")
            cur.execute("DELETE FROM ingresos")
            cur.execute("DELETE FROM devices")

            cur.execute(
                """
                INSERT INTO devices(
                    customer_id, marca_id, model_id, numero_serie, numero_interno,
                    n_de_control, tipo_equipo, variante, alquilado, alquiler_a, ubicacion_id, mg_estado
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE,NULL,%s,'activo')
                """,
                [
                    self.customer_cli,
                    self.marca_id,
                    self.model_id,
                    "N5143421",
                    "MG 1001",
                    "MG 1001",
                    "Concentrador",
                    "",
                    self.loc_taller,
                ],
            )
            self.device_id = self._last_id(cur)
            cur.execute(
                """
                INSERT INTO ingresos(
                    device_id, estado, motivo, fecha_ingreso, fecha_creacion,
                    ubicacion_id, alquilado, alquiler_a, alquiler_remito, alquiler_fecha
                )
                VALUES (%s,%s,%s,%s,%s,%s,FALSE,NULL,NULL,NULL)
                """,
                [
                    self.device_id,
                    "reparado",
                    "reparacion",
                    timezone.make_aware(datetime(2025, 6, 10, 12, 0, 0)),
                    timezone.make_aware(datetime(2025, 6, 10, 12, 0, 0)),
                    self.loc_taller,
                ],
            )
            self.ingreso_id = self._last_id(cur)

    def _write_novamed_workbook(self, path: Path):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "VENTAS NOVAMED - MG BIO SA"
        ws.append(["Cód. Gen.", "Fecha FC", "Comp.", "Equipo", "N/S", "Precio Neto", "IVA", "Comp. - Nro. Remito"])
        ws.append(["1115007", "2025-06-15", "FC A 0001", "Concentrador AirSep", "N5143421", 1, 0.21, "R-1"])
        ws2 = workbook.create_sheet("VENTAS NOVAMED - SEPID SA")
        ws2.append(["Cód. Gen.", "Fecha FC", "Comp.", "Equipo", "N/S", "Precio Neto", "IVA", "Comp. - Nro. Remito"])
        ws3 = workbook.create_sheet("ALQUILERES NOVAMED - SEPID SA")
        ws3.append(["Cód. Gen.", "Equipo", "N/S", "Precio Neto", "IVA"])
        ws4 = workbook.create_sheet("PRÉSTAMO")
        ws4.append(["Cód. Gen.", "Equipo", "N/S", "ID"])
        workbook.save(path)

    def test_trace_pending_devices_generates_sale_proposal(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            workbook_path = tmp_path / "novamed.xlsx"
            access_path = tmp_path / "dummy.accdb"
            access_path.write_text("", encoding="utf-8")
            self._write_novamed_workbook(workbook_path)
            out_dir = tmp_path / "out"

            with patch("service.pending_device_trace.load_access_rows", return_value=({}, None)):
                with patch("service.pending_device_trace.BejermanLookup.lookup_candidate", return_value=([], [])):
                    call_command(
                        "trace_pending_devices",
                        "--cutoff-date",
                        "2026-01-01",
                        "--xlsx-novamed",
                        str(workbook_path),
                        "--access-db",
                        str(access_path),
                        "--out-dir",
                        str(out_dir),
                    )

            with (out_dir / "propuestas.csv").open("r", encoding="utf-8", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(int(row["ingreso_id"]), self.ingreso_id)
        self.assertEqual(row["action"], "vendido_entregado")
        self.assertEqual(row["confidence"], "alta")
        payload = json.loads(row["payload_json"])
        self.assertEqual(payload["venta_customer_nombre"], "NOVAMED SA")
        self.assertEqual(payload["factura_numero"], "FC A 0001")

    def test_apply_traced_device_corrections_marks_sale_and_skips_unapproved(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            csv_path = tmp_path / "propuestas.csv"
            out_dir = tmp_path / "applied"
            rows = [
                {
                    "ingreso_id": self.ingreso_id,
                    "device_id": self.device_id,
                    "approved": 1,
                    "action": "vendido_entregado",
                    "fecha_venta": "2025-06-15 00:00:00-03:00",
                    "fecha_efectiva": "2025-06-15 00:00:00-03:00",
                    "venta_customer_id": self.customer_novamed,
                    "venta_customer_nombre": "NOVAMED SA",
                    "venta_numero_alternativo": "MG 1001",
                    "factura_numero": "FC A 0001",
                    "remito_numero": "R-1",
                    "payload_json": "",
                },
                {
                    "ingreso_id": self.ingreso_id,
                    "device_id": self.device_id,
                    "approved": 0,
                    "action": "baja",
                    "payload_json": "",
                },
            ]
            with csv_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                for row in rows:
                    writer.writerow(row)

            call_command(
                "apply_traced_device_corrections",
                "--input",
                str(csv_path),
                "--actor-email",
                self.actor.email,
                "--out-dir",
                str(out_dir),
            )

            with connection.cursor() as cur:
                cur.execute("SELECT estado, factura_numero, remito_salida FROM ingresos WHERE id=%s", [self.ingreso_id])
                ingreso = cur.fetchone()
                cur.execute("SELECT mg_estado, mg_venta_factura_numero, mg_venta_remito_numero FROM devices WHERE id=%s", [self.device_id])
                device = cur.fetchone()
                cur.execute("SELECT COUNT(*) FROM ingreso_events WHERE ticket_id=%s AND a_estado='vendido_entregado'", [self.ingreso_id])
                ingreso_events = int(cur.fetchone()[0])
                cur.execute("SELECT COUNT(*) FROM device_mg_events WHERE device_id=%s AND accion='venta'", [self.device_id])
                mg_events = int(cur.fetchone()[0])

            with (out_dir / "applied_results.csv").open("r", encoding="utf-8", newline="") as handle:
                result_rows = list(csv.DictReader(handle))

        self.assertEqual(len(result_rows), 2)
        self.assertEqual(result_rows[0]["status"], "applied", result_rows[0]["detail"])
        self.assertEqual(result_rows[1]["status"], "skipped")
        self.assertEqual(ingreso[0], "vendido_entregado")
        self.assertEqual(ingreso[1], "FC A 0001")
        self.assertEqual(ingreso[2], "R-1")
        self.assertEqual(device[0], "inactivo_venta")
        self.assertEqual(device[1], "FC A 0001")
        self.assertEqual(device[2], "R-1")
        self.assertEqual(ingreso_events, 1)
        self.assertEqual(mg_events, 1)
