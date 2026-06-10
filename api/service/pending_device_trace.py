from __future__ import annotations

import csv
import json
import re
import subprocess
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

from django.db import connection, transaction
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime

from .bejerman_sync import BejermanConfigError, BejermanSDKClient, BejermanTransientError, _records_for_partida
from .models import User

DEFAULT_CUTOFF_DATE = "2026-01-01"
DEFAULT_XLSX_NOVAMED = (
    r"Z:\Servicio Tecnico\Registros\Aporte de Equipamiento - SEPID SA x NOVAMED SA - Al 09.04.2026.V2.xlsx"
)
DEFAULT_ACCESS_DB = r"Z:\Servicio Tecnico\1_SISTEMA REPARACIONES\2025-06\Tablas2025 MG-SEPID 2.0.accdb"
DEFAULT_SALE_CUSTOMER_NAME = "NOVAMED SA"
DEFAULT_SALE_REASON = "Regularización histórica por venta confirmada en registro NOVAMED."
DEFAULT_RENT_REASON = "Regularización histórica por alquiler confirmado en fuentes externas."
DEFAULT_BAJA_REASON = "Regularización histórica por baja confirmada en trazas internas."
TERMINAL_STATES = (
    "entregado",
    "alquilado",
    "baja",
    "vendido_pendiente_entrega",
    "vendido_entregado",
)
WORKBOOK_NAME_MAP = {
    "VENTAS NOVAMED - MG BIO SA": "ventas_mg",
    "VENTAS NOVAMED - SEPID SA": "ventas_sepid",
    "ALQUILERES NOVAMED - SEPID SA": "alquileres_sepid",
    "PRESTAMO": "prestamo",
}
WORKBOOK_SHEETS = (
    "ventas_mg",
    "ventas_sepid",
    "alquileres_sepid",
    "prestamo",
)
BOOL_TRUE = {"1", "true", "yes", "y", "si", "sí", "on", "x"}
BEJERMAN_DEPOSIT_INTERPRETATIONS = {
    "STR": {
        "classification": "stock_servicio",
        "conclusion": "sin_indicio_salida",
        "detail": "Depósito STR: el equipo sigue en stock/servicio y no acredita venta ni alquiler.",
    },
    "STL": {
        "classification": "stock_alquiler_mg",
        "conclusion": "indicio_alquiler",
        "detail": "Depósito STL: stock interno MG para alquiler; es indicio de alquiler, no de venta.",
    },
    "STC": {
        "classification": "stock_cliente",
        "conclusion": "indicio_entrega_cliente",
        "detail": "Depósito STC: stock de cliente/listo para salida; sugiere entrega, pero no define venta o devolución.",
    },
    "STCL": {
        "classification": "stock_cliente",
        "conclusion": "indicio_entrega_cliente",
        "detail": "Depósito STCL: stock de cliente/listo para salida; sugiere entrega, pero no define venta o devolución.",
    },
    "SALIDA": {
        "classification": "salida_registrada",
        "conclusion": "indicio_salida_final",
        "detail": "Depósito SALIDA: Bejerman registra una salida final, pero por sí solo no distingue venta de devolución.",
    },
}


def _norm_text(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFKD", raw)
    cleaned = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", cleaned).strip()


def _norm_upper_ascii(value: Any) -> str:
    return _norm_text(value).upper()


def _norm_sheet_name(value: Any) -> str:
    return _norm_upper_ascii(value)


def _norm_ns(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", _norm_upper_ascii(value))


def _norm_stockcode(value: Any) -> Optional[str]:
    if value is None:
        return None
    raw = _norm_upper_ascii(value)
    if not raw:
        return None
    match = re.match(r"^(MG|NM|NV|CE)\s*(\d{1,4})$", raw)
    if not match:
        match = re.match(r"^(MG|NM|NV|CE)[^0-9]*(\d{1,4})$", raw)
    if not match:
        return None
    prefix, number = match.group(1), match.group(2)
    return f"{prefix} {number.zfill(4)}"


def _bejerman_deposit_info(deposit: Any) -> dict[str, str]:
    deposit_code = _norm_upper_ascii(deposit)
    info = BEJERMAN_DEPOSIT_INTERPRETATIONS.get(deposit_code)
    if info:
        return {
            "deposit_code": deposit_code,
            "classification": info["classification"],
            "conclusion": info["conclusion"],
            "detail": info["detail"],
        }
    if deposit_code:
        return {
            "deposit_code": deposit_code,
            "classification": "deposito_no_mapeado",
            "conclusion": "indeterminado",
            "detail": f"Depósito {deposit_code}: sin interpretación automática; requiere revisión manual.",
        }
    return {
        "deposit_code": "",
        "classification": "sin_deposito",
        "conclusion": "indeterminado",
        "detail": "Bejerman devolvió la partida sin identificar depósito.",
    }


def _parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def _parse_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return default
    return text in BOOL_TRUE


def _parse_datetime_or_date(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, 0, 0, 0)
    text = str(value).strip()
    if not text or text == "-":
        return None
    normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
    parsed = parse_datetime(normalized) or parse_datetime(text)
    if parsed:
        return parsed
    parsed_date = parse_date(text)
    if parsed_date:
        return datetime(parsed_date.year, parsed_date.month, parsed_date.day, 0, 0, 0)
    for sep in ("/", "-", "."):
        parts = text.split(sep)
        if len(parts) != 3 or not all(part.isdigit() for part in parts):
            continue
        if len(parts[0]) == 4:
            year, month, day = parts
        elif len(parts[2]) == 4:
            day, month, year = parts
        else:
            continue
        try:
            return datetime(int(year), int(month), int(day), 0, 0, 0)
        except ValueError:
            return None
    return None


def _parse_effective_datetime(value: Any) -> Optional[datetime]:
    parsed = _parse_datetime_or_date(value)
    if not parsed:
        return None
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _format_scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (dict, list)):
        return _json_dumps(value)
    return value


def _order_key_from_row(row: dict[str, Any], *field_names: str) -> tuple[int, str]:
    for field_name in field_names:
        raw = row.get(field_name)
        parsed = _parse_effective_datetime(raw)
        if parsed:
            return (1, parsed.isoformat())
    return (0, "")


def _table_exists(table_name: str, schema: Optional[str] = None) -> bool:
    try:
        if connection.vendor == "postgresql":
            if schema:
                sql = """
                    SELECT 1
                      FROM information_schema.tables
                     WHERE table_schema = %s
                       AND table_name = %s
                     LIMIT 1
                """
                params = [schema, table_name]
            else:
                sql = """
                    SELECT 1
                      FROM information_schema.tables
                     WHERE table_name = %s
                       AND table_schema = ANY(current_schemas(true))
                     LIMIT 1
                """
                params = [table_name]
        else:
            sql = """
                SELECT 1
                  FROM information_schema.tables
                 WHERE table_name = %s
                 LIMIT 1
            """
            params = [table_name]
        with connection.cursor() as cur:
            cur.execute(sql, params)
            return bool(cur.fetchone())
    except Exception:
        return False


def _table_has_column(table_name: str, column_name: str, schema: Optional[str] = None) -> bool:
    try:
        if connection.vendor == "postgresql":
            if schema:
                sql = """
                    SELECT 1
                      FROM information_schema.columns
                     WHERE table_schema = %s
                       AND table_name = %s
                       AND column_name = %s
                     LIMIT 1
                """
                params = [schema, table_name, column_name]
            else:
                sql = """
                    SELECT 1
                      FROM information_schema.columns
                     WHERE table_name = %s
                       AND column_name = %s
                       AND table_schema = ANY(current_schemas(true))
                     LIMIT 1
                """
                params = [table_name, column_name]
        else:
            sql = """
                SELECT 1
                  FROM information_schema.columns
                 WHERE table_name = %s
                   AND column_name = %s
                 LIMIT 1
            """
            params = [table_name, column_name]
        with connection.cursor() as cur:
            cur.execute(sql, params)
            return bool(cur.fetchone())
    except Exception:
        return False


def _fetchall_dicts(sql: str, params: Optional[Sequence[Any]] = None) -> list[dict[str, Any]]:
    with connection.cursor() as cur:
        cur.execute(sql, params or [])
        columns = [col[0] for col in cur.description]
        return [dict(zip(columns, row)) for row in cur.fetchall()]


def _fetchone_dict(sql: str, params: Optional[Sequence[Any]] = None) -> Optional[dict[str, Any]]:
    rows = _fetchall_dicts(sql, params)
    return rows[0] if rows else None


def _exec(sql: str, params: Optional[Sequence[Any]] = None) -> int:
    with connection.cursor() as cur:
        cur.execute(sql, params or [])
        return int(cur.rowcount or 0)


def _set_audit_actor(user_id: Optional[int], role: str = "") -> None:
    if not user_id:
        return
    with connection.cursor() as cur:
        cur.execute("SET app.user_id = %s;", [str(user_id)])
        cur.execute("SET app.user_role = %s;", [role or ""])


def _dash_location_id() -> Optional[int]:
    row = _fetchone_dict("SELECT id FROM locations WHERE nombre='-' LIMIT 1")
    return _parse_int(row.get("id")) if row else None


def _taller_location_id() -> Optional[int]:
    row = _fetchone_dict("SELECT id FROM locations WHERE LOWER(nombre)=LOWER(%s) LIMIT 1", ["Taller"])
    return _parse_int(row.get("id")) if row else None


def _resolve_actor(actor_email: str = "") -> tuple[Optional[int], str]:
    email = str(actor_email or "").strip()
    if not email:
        return None, ""
    user = User.objects.filter(email__iexact=email, activo=True).first()
    if not user:
        raise ValueError(f"Usuario no encontrado o inactivo: {email}")
    return int(user.id), str(user.rol or "")


def _mg_owner_customer_id() -> Optional[int]:
    row = _fetchone_dict(
        "SELECT id FROM customers WHERE LOWER(razon_social) LIKE %s ORDER BY id ASC LIMIT 1",
        ["%mg%bio%"],
    )
    return _parse_int(row.get("id")) if row else None


def _resolve_sale_customer_id(preferred_name: str = DEFAULT_SALE_CUSTOMER_NAME) -> Optional[int]:
    exact = _fetchone_dict(
        "SELECT id FROM customers WHERE LOWER(razon_social)=LOWER(%s) ORDER BY id ASC LIMIT 1",
        [preferred_name],
    )
    if exact:
        return _parse_int(exact.get("id"))
    fuzzy = _fetchone_dict(
        "SELECT id FROM customers WHERE LOWER(razon_social) LIKE %s ORDER BY id ASC LIMIT 1",
        [f"%{preferred_name.lower()}%"],
    )
    return _parse_int(fuzzy.get("id")) if fuzzy else None


def _is_airsep_family(candidate: dict[str, Any]) -> bool:
    haystack = " ".join(
        [
            str(candidate.get("marca") or ""),
            str(candidate.get("modelo") or ""),
            str(candidate.get("tipo_equipo") or ""),
            str(candidate.get("variante") or ""),
        ]
    )
    text = _norm_upper_ascii(haystack)
    return any(token in text for token in ("AIR SEP", "AIRSEP", "NEW LIFE", "NEWLIFE"))


def serial_search_variants(candidate: dict[str, Any]) -> list[dict[str, str]]:
    serial_raw = str(candidate.get("numero_serie") or "").strip()
    serial_norm = _norm_ns(serial_raw)
    if not serial_norm:
        return []
    variants: list[dict[str, str]] = [
        {
            "value": serial_norm,
            "match_type": "serial_exact",
            "serial_busqueda": serial_raw or serial_norm,
        }
    ]
    is_airsep = _is_airsep_family(candidate)
    if serial_norm.startswith("N") and len(serial_norm) > 1:
        variants.append(
            {
                "value": serial_norm[1:],
                "match_type": "serial_variant_sin_n",
                "serial_busqueda": serial_norm[1:],
            }
        )
    elif is_airsep:
        variants.append(
            {
                "value": f"N{serial_norm}",
                "match_type": "serial_variant_con_n",
                "serial_busqueda": f"N{serial_norm}",
            }
        )
    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in variants:
        key = item["value"]
        if key in seen:
            continue
        deduped.append(item)
        seen.add(key)
    return deduped


@dataclass
class WorkbookRecord:
    sheet_key: str
    source_name: str
    row_number: int
    codigo_gen: str
    fecha: Optional[datetime]
    comprobante: str
    equipo: str
    serial_raw: str
    serial_norm: str
    precio_neto: str
    iva: str
    remito: str


@dataclass
class WorkbookIndex:
    by_sheet: dict[str, dict[str, list[WorkbookRecord]]]

    def lookup(self, sheet_key: str, serial_norm: str) -> list[WorkbookRecord]:
        return list(self.by_sheet.get(sheet_key, {}).get(serial_norm, []))


def load_novamed_workbook(path: str) -> WorkbookIndex:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        resolved: dict[str, Any] = {}
        sheet_map = {_norm_sheet_name(name): name for name in workbook.sheetnames}
        for raw_name, sheet_key in WORKBOOK_NAME_MAP.items():
            real_name = sheet_map.get(raw_name)
            if not real_name:
                raise ValueError(f"No se encontró la hoja requerida: {raw_name}")
            resolved[sheet_key] = workbook[real_name]

        by_sheet: dict[str, dict[str, list[WorkbookRecord]]] = {
            sheet_key: defaultdict(list) for sheet_key in WORKBOOK_SHEETS
        }
        for sheet_key, ws in resolved.items():
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if sheet_key in ("ventas_mg", "ventas_sepid"):
                    codigo_gen = str(row[0] or "").strip() if len(row) > 0 else ""
                    fecha = _parse_datetime_or_date(row[1] if len(row) > 1 else None)
                    comprobante = str(row[2] or "").strip() if len(row) > 2 else ""
                    equipo = str(row[3] or "").strip() if len(row) > 3 else ""
                    serial_raw = str(row[4] or "").strip() if len(row) > 4 else ""
                    precio_neto = str(row[5] or "").strip() if len(row) > 5 else ""
                    iva = str(row[6] or "").strip() if len(row) > 6 else ""
                    remito = str(row[7] or "").strip() if len(row) > 7 else ""
                else:
                    codigo_gen = str(row[0] or "").strip() if len(row) > 0 else ""
                    fecha = None
                    comprobante = ""
                    equipo = str(row[1] or "").strip() if len(row) > 1 else ""
                    serial_raw = str(row[2] or "").strip() if len(row) > 2 else ""
                    precio_neto = str(row[3] or "").strip() if len(row) > 3 else ""
                    iva = str(row[4] or "").strip() if len(row) > 4 else ""
                    remito = ""
                serial_norm = _norm_ns(serial_raw)
                if not serial_norm:
                    continue
                record = WorkbookRecord(
                    sheet_key=sheet_key,
                    source_name=sheet_key,
                    row_number=row_number,
                    codigo_gen=codigo_gen,
                    fecha=fecha,
                    comprobante=comprobante,
                    equipo=equipo,
                    serial_raw=serial_raw,
                    serial_norm=serial_norm,
                    precio_neto=precio_neto,
                    iva=iva,
                    remito=remito,
                )
                by_sheet[sheet_key][serial_norm].append(record)
        return WorkbookIndex(by_sheet=by_sheet)
    finally:
        workbook.close()


class BejermanLookup:
    def __init__(self):
        self._client: Optional[BejermanSDKClient] = None
        self._init_error: Optional[str] = None
        self._cache: dict[str, tuple[list[dict[str, Any]], Optional[str]]] = {}

    def _client_or_error(self) -> tuple[Optional[BejermanSDKClient], Optional[str]]:
        if self._client or self._init_error:
            return self._client, self._init_error
        try:
            self._client = BejermanSDKClient()
            if not self._client.wsdl_url:
                raise BejermanConfigError("BEJERMAN_WSDL_URL requerido")
        except Exception as exc:
            self._init_error = str(exc)
        return self._client, self._init_error

    @staticmethod
    def _pick_first(record: dict[str, Any], *candidates: str) -> Any:
        lowered = {str(key).lower(): value for key, value in record.items()}
        for candidate in candidates:
            if candidate in record:
                return record.get(candidate)
            lowered_candidate = candidate.lower()
            if lowered_candidate in lowered:
                return lowered[lowered_candidate]
        return None

    def query_variant(self, serial_variant: str) -> tuple[list[dict[str, Any]], Optional[str]]:
        cached = self._cache.get(serial_variant)
        if cached is not None:
            return cached
        client, init_error = self._client_or_error()
        if init_error or client is None:
            result = ([], init_error or "Cliente Bejerman no disponible")
            self._cache[serial_variant] = result
            return result
        try:
            response = client.stock_by_deposit_partida("STR", serial_variant)
            rows = _records_for_partida(response, serial_variant)
            result = (rows, None)
        except (BejermanConfigError, BejermanTransientError) as exc:
            result = ([], str(exc))
        except Exception as exc:
            result = ([], str(exc))
        self._cache[serial_variant] = result
        return result

    def lookup_candidate(self, candidate: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        evidences: list[dict[str, Any]] = []
        conflicts: list[dict[str, Any]] = []
        seen_rows: set[tuple[str, str, str]] = set()
        for serial_info in serial_search_variants(candidate):
            variant = serial_info["value"]
            rows, error = self.query_variant(variant)
            if error:
                conflicts.append(
                    {
                        "ingreso_id": candidate["ingreso_id"],
                        "device_id": candidate["device_id"],
                        "conflict_type": "bejerman_error",
                        "serial_busqueda": serial_info["serial_busqueda"],
                        "serial_encontrado": "",
                        "os_access": "",
                        "detalle": error,
                        "payload_json": "",
                    }
                )
                continue
            for record in rows:
                serial_found = str(
                    self._pick_first(
                        record,
                        "Comprobante_ArtPartida",
                        "Art_Partida",
                        "ArtPartida",
                        "Partida",
                        "Item_Partida",
                        "Serie",
                        "NumeroSerie",
                    )
                    or ""
                ).strip()
                deposit = str(
                    self._pick_first(
                        record,
                        "Comprobante_ArtDeposito",
                        "Art_CodDeposito",
                        "ArtDeposito",
                        "Deposito",
                        "Item_Deposito",
                    )
                    or ""
                ).strip()
                article_code = str(
                    self._pick_first(
                        record,
                        "Comprobante_Art_CodGen",
                        "Art_CodGenerico",
                        "Art_CodGen",
                        "Art_CodReducido",
                        "CodigoArticulo",
                        "Codigo",
                    )
                    or ""
                ).strip()
                article_description = str(
                    self._pick_first(
                        record,
                        "Art_DescripcionGeneral",
                        "Art_DescripcionReducida",
                        "Descripcion",
                        "Nombre",
                    )
                    or ""
                ).strip()
                row_key = (serial_found, deposit, article_code)
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)
                deposit_info = _bejerman_deposit_info(deposit)
                evidences.append(
                    {
                        "ingreso_id": candidate["ingreso_id"],
                        "device_id": candidate["device_id"],
                        "source": "bejerman",
                        "source_priority": 2,
                        "match_type": serial_info["match_type"],
                        "serial_busqueda": serial_info["serial_busqueda"],
                        "serial_encontrado": serial_found or variant,
                        "fecha_fuente": "",
                        "factura": "",
                        "remito": "",
                        "deposito": deposit,
                        "os_access": "",
                        "confidence_hint": "complementaria",
                        "classification": deposit_info["classification"],
                        "conclusion_operativa": deposit_info["conclusion"],
                        "detalle": (
                            f"Artículo {article_code or '-'} {article_description or ''} | {deposit_info['detail']}"
                        ).strip(),
                    }
                )
        return evidences, conflicts


def _access_select_prefix() -> str:
    return (
        "SELECT Id, "
        "IIF([Fecha Ingreso] IS NULL, NULL, Format([Fecha Ingreso], 'yyyy-mm-dd hh:nn:ss')) AS FechaIngreso, "
        "NumeroSerie, NdeControl, Entregado, Alquilado, "
        "IIF([FechaEntrega] IS NULL, NULL, Format([FechaEntrega], 'yyyy-mm-dd hh:nn:ss')) AS FechaEntrega, "
        "Venta, Estado, Remito, Factura, Marca, Modelo, Recibido, Comentarios, RecibeAlquiler, CargoAlquiler, "
        "RemitoIngreso "
        "FROM Servicio "
    )


def _quote_access_text(value: Any) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _build_access_query(ids: Sequence[int]) -> str:
    id_list = ",".join(str(int(item)) for item in ids)
    return _access_select_prefix() + f"WHERE Id IN ({id_list}) ORDER BY Id"


def _build_access_fallback_query(remitos: Sequence[int], mg_codes: Sequence[str]) -> str:
    clauses: list[str] = []
    if remitos:
        remito_list = ",".join(str(int(item)) for item in remitos)
        clauses.append(f"(RemitoIngreso IN ({remito_list}) OR Remito IN ({remito_list}))")
    if mg_codes:
        mg_list = ",".join(_quote_access_text(item) for item in mg_codes)
        clauses.append(f"(NumeroSerie IN ({mg_list}) OR NdeControl IN ({mg_list}))")
    if not clauses:
        raise ValueError("Se requieren remitos o códigos MG para la búsqueda fallback en Access.")
    return _access_select_prefix() + f"WHERE {' OR '.join(clauses)} ORDER BY Id"


def _run_access_query(access_db: str, query: str) -> list[dict[str, Any]]:
    path = Path(access_db)
    if not path.exists():
        raise FileNotFoundError(f"No existe la base Access: {path}")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
        temp_csv = Path(tmp.name)
    ps_script = f"""
$ErrorActionPreference = 'Stop'
Add-Type -AssemblyName System.Data
$dbPath = '{str(path).replace("'", "''")}'
$outPath = '{str(temp_csv).replace("'", "''")}'
$connection = $null
try {{
  $connection = New-Object System.Data.OleDb.OleDbConnection("Provider=Microsoft.ACE.OLEDB.12.0;Data Source=$dbPath;Persist Security Info=False;")
  $connection.Open()
}} catch {{
  $connection = New-Object System.Data.OleDb.OleDbConnection("Provider=Microsoft.ACE.OLEDB.16.0;Data Source=$dbPath;Persist Security Info=False;")
  $connection.Open()
}}
$command = $connection.CreateCommand()
$command.CommandText = @'
{query}
'@
$adapter = New-Object System.Data.OleDb.OleDbDataAdapter $command
$table = New-Object System.Data.DataTable
[void]$adapter.Fill($table)
$connection.Close()
$table | Export-Csv -Path $outPath -NoTypeInformation -Encoding UTF8
"""
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            check=True,
            capture_output=True,
            text=True,
        )
        rows: list[dict[str, Any]] = []
        with temp_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                rows.append(dict(row))
        return rows
    finally:
        temp_csv.unlink(missing_ok=True)


def load_access_rows(access_db: str, ids: Sequence[int]) -> tuple[dict[int, dict[str, Any]], Optional[str]]:
    if not ids:
        return {}, None
    path = Path(access_db)
    if not path.exists():
        return {}, f"No existe la base Access: {path}"
    rows: dict[int, dict[str, Any]] = {}
    try:
        chunk_size = 150
        for start in range(0, len(ids), chunk_size):
            chunk = ids[start : start + chunk_size]
            for row in _run_access_query(access_db, _build_access_query(chunk)):
                row_id = _parse_int(row.get("Id"))
                if row_id is None:
                    continue
                rows[row_id] = row
    except Exception as exc:
        return {}, str(exc)
    return rows, None


def load_access_fallback_rows(
    access_db: str,
    *,
    remitos: Sequence[int],
    mg_codes: Sequence[str],
) -> tuple[list[dict[str, Any]], Optional[str]]:
    clean_remitos = sorted({int(item) for item in remitos if item is not None})
    clean_mg_codes = sorted({str(item).strip() for item in mg_codes if str(item).strip()})
    if not clean_remitos and not clean_mg_codes:
        return [], None
    try:
        chunk_size = 120
        rows: list[dict[str, Any]] = []
        seen_keys: set[tuple[str, str, str, str]] = set()
        max_chunks = max(
            (len(clean_remitos) + chunk_size - 1) // chunk_size if clean_remitos else 0,
            (len(clean_mg_codes) + chunk_size - 1) // chunk_size if clean_mg_codes else 0,
            1,
        )
        for index in range(max_chunks):
            remito_chunk = clean_remitos[index * chunk_size : (index + 1) * chunk_size]
            mg_chunk = clean_mg_codes[index * chunk_size : (index + 1) * chunk_size]
            if not remito_chunk and not mg_chunk:
                continue
            query = _build_access_fallback_query(remito_chunk, mg_chunk)
            for row in _run_access_query(access_db, query):
                row_key = (
                    str(row.get("Id") or ""),
                    str(row.get("NumeroSerie") or ""),
                    str(row.get("NdeControl") or ""),
                    str(row.get("RemitoIngreso") or ""),
                )
                if row_key in seen_keys:
                    continue
                seen_keys.add(row_key)
                rows.append(row)
        return rows, None
    except Exception as exc:
        return [], str(exc)


def _build_candidate_sql(ids: Sequence[int] | None = None) -> tuple[str, list[Any]]:
    has_n_de_control = _table_has_column("devices", "n_de_control")
    has_tipo = _table_has_column("devices", "tipo_equipo")
    has_variante = _table_has_column("devices", "variante")
    has_model_tipo = _table_has_column("models", "tipo_equipo")
    has_model_variante = _table_has_column("models", "variante")
    has_mg_estado = _table_has_column("devices", "mg_estado")
    has_mg_sale_fecha = _table_has_column("devices", "mg_venta_fecha")
    has_mg_sale_factura = _table_has_column("devices", "mg_venta_factura_numero")
    has_mg_sale_remito = _table_has_column("devices", "mg_venta_remito_numero")
    has_mg_sale_customer = _table_has_column("devices", "mg_venta_customer_id")
    has_mg_sale_alt = _table_has_column("devices", "mg_venta_numero_alternativo")
    has_alquiler_a = _table_has_column("devices", "alquiler_a")
    has_ubicacion = _table_has_column("devices", "ubicacion_id")
    has_remito_ingreso = _table_has_column("ingresos", "remito_ingreso")

    select_parts = [
        "i.id AS ingreso_id",
        "d.id AS device_id",
        "d.customer_id",
        "COALESCE(c.razon_social,'') AS cliente",
        "COALESCE(b.nombre,'') AS marca",
        "COALESCE(m.nombre,'') AS modelo",
        ("COALESCE(d.n_de_control,'') AS n_de_control" if has_n_de_control else "'' AS n_de_control"),
        (
            "COALESCE(NULLIF(d.tipo_equipo,''), NULLIF(m.tipo_equipo,''), '') AS tipo_equipo"
            if has_tipo and has_model_tipo
            else (
                "COALESCE(NULLIF(d.tipo_equipo,''), '') AS tipo_equipo"
                if has_tipo
                else ("COALESCE(NULLIF(m.tipo_equipo,''), '') AS tipo_equipo" if has_model_tipo else "'' AS tipo_equipo")
            )
        ),
        (
            "COALESCE(NULLIF(d.variante,''), NULLIF(m.variante,''), '') AS variante"
            if has_variante and has_model_variante
            else (
                "COALESCE(NULLIF(d.variante,''), '') AS variante"
                if has_variante
                else ("COALESCE(NULLIF(m.variante,''), '') AS variante" if has_model_variante else "'' AS variante")
            )
        ),
        "COALESCE(d.numero_serie,'') AS numero_serie",
        "COALESCE(d.numero_interno,'') AS numero_interno",
        ("COALESCE(i.remito_ingreso,'') AS remito_ingreso" if has_remito_ingreso else "'' AS remito_ingreso"),
        "COALESCE(i.estado::text,'') AS estado_actual",
        "COALESCE(i.fecha_ingreso, i.fecha_creacion) AS fecha_base",
        "COALESCE(i.alquilado, FALSE) AS ingreso_alquilado",
        "COALESCE(d.alquilado, FALSE) AS device_alquilado",
        ("COALESCE(d.alquiler_a,'') AS alquiler_a" if has_alquiler_a else "'' AS alquiler_a"),
        ("COALESCE(loc.nombre,'') AS ubicacion_nombre" if has_ubicacion else "'' AS ubicacion_nombre"),
        ("COALESCE(d.mg_estado,'activo') AS mg_estado" if has_mg_estado else "'activo' AS mg_estado"),
        ("d.mg_venta_fecha AS mg_venta_fecha" if has_mg_sale_fecha else "NULL AS mg_venta_fecha"),
        (
            "COALESCE(d.mg_venta_factura_numero,'') AS mg_venta_factura_numero"
            if has_mg_sale_factura
            else "'' AS mg_venta_factura_numero"
        ),
        (
            "COALESCE(d.mg_venta_remito_numero,'') AS mg_venta_remito_numero"
            if has_mg_sale_remito
            else "'' AS mg_venta_remito_numero"
        ),
        ("d.mg_venta_customer_id AS mg_venta_customer_id" if has_mg_sale_customer else "NULL AS mg_venta_customer_id"),
        (
            "COALESCE(csale.razon_social,'') AS mg_venta_customer_nombre"
            if has_mg_sale_customer
            else "'' AS mg_venta_customer_nombre"
        ),
        (
            "COALESCE(d.mg_venta_numero_alternativo,'') AS mg_venta_numero_alternativo"
            if has_mg_sale_alt
            else "'' AS mg_venta_numero_alternativo"
        ),
    ]
    joins = [
        "JOIN devices d ON d.id = i.device_id",
        "LEFT JOIN customers c ON c.id = d.customer_id",
        "LEFT JOIN marcas b ON b.id = d.marca_id",
        "LEFT JOIN models m ON m.id = d.model_id",
    ]
    if has_ubicacion:
        joins.append("LEFT JOIN locations loc ON loc.id = i.ubicacion_id")
    if has_mg_sale_customer:
        joins.append("LEFT JOIN customers csale ON csale.id = d.mg_venta_customer_id")

    sql = f"""
        SELECT {", ".join(select_parts)}
          FROM ingresos i
          {' '.join(joins)}
         WHERE DATE(COALESCE(i.fecha_ingreso, i.fecha_creacion)) < %s
           AND COALESCE(i.estado::text, '') NOT IN ({", ".join(["%s"] * len(TERMINAL_STATES))})
    """
    params: list[Any] = [DEFAULT_CUTOFF_DATE, *TERMINAL_STATES]
    if ids:
        sql += " AND i.id = ANY(%s)"
        params.append(list(ids))
    sql += " ORDER BY COALESCE(i.fecha_ingreso, i.fecha_creacion) ASC, i.id ASC"
    return sql, params


def fetch_candidates(cutoff_date: str, ids: Sequence[int] | None = None) -> list[dict[str, Any]]:
    sql, params = _build_candidate_sql(ids)
    params[0] = cutoff_date
    return _fetchall_dicts(sql, params)


def build_duplicate_serial_map(candidates: Sequence[dict[str, Any]]) -> dict[str, list[int]]:
    serial_map: dict[str, list[int]] = defaultdict(list)
    for candidate in candidates:
        serial_norm = _norm_ns(candidate.get("numero_serie"))
        if not serial_norm:
            continue
        serial_map[serial_norm].append(int(candidate["ingreso_id"]))
    return {key: value for key, value in serial_map.items() if len(value) > 1}


def _build_workbook_evidence(
    candidate: dict[str, Any],
    serial_info: dict[str, str],
    record: WorkbookRecord,
    confidence_hint: str,
) -> dict[str, Any]:
    return {
        "ingreso_id": candidate["ingreso_id"],
        "device_id": candidate["device_id"],
        "source": f"excel_{record.sheet_key}",
        "source_priority": 1,
        "match_type": serial_info["match_type"],
        "serial_busqueda": serial_info["serial_busqueda"],
        "serial_encontrado": record.serial_raw,
        "fecha_fuente": _format_scalar(record.fecha),
        "factura": record.comprobante,
        "remito": record.remito,
        "deposito": "",
        "os_access": "",
        "confidence_hint": confidence_hint,
        "detalle": f"{record.equipo} (fila {record.row_number})".strip(),
    }


def collect_workbook_matches(candidate: dict[str, Any], workbook_index: WorkbookIndex) -> dict[str, Any]:
    sale_matches: list[dict[str, Any]] = []
    rental_matches: list[dict[str, Any]] = []
    loan_matches: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    candidate_serial_norm = _norm_ns(candidate.get("numero_serie"))
    for serial_info in serial_search_variants(candidate):
        for record in workbook_index.lookup("ventas_mg", serial_info["value"]):
            sale_matches.append({"record": record, "serial_info": serial_info})
        for record in workbook_index.lookup("ventas_sepid", serial_info["value"]):
            sale_matches.append({"record": record, "serial_info": serial_info})
        for record in workbook_index.lookup("alquileres_sepid", serial_info["value"]):
            rental_matches.append({"record": record, "serial_info": serial_info})
        for record in workbook_index.lookup("prestamo", serial_info["value"]):
            loan_matches.append({"record": record, "serial_info": serial_info})

    if len(sale_matches) > 1:
        conflicts.append(
            {
                "ingreso_id": candidate["ingreso_id"],
                "device_id": candidate["device_id"],
                "conflict_type": "excel_venta_multiple",
                "serial_busqueda": candidate.get("numero_serie") or "",
                "serial_encontrado": candidate_serial_norm,
                "os_access": "",
                "detalle": f"Se encontraron {len(sale_matches)} coincidencias de venta en Excel.",
                "payload_json": "",
            }
        )
    if len(rental_matches) > 1:
        conflicts.append(
            {
                "ingreso_id": candidate["ingreso_id"],
                "device_id": candidate["device_id"],
                "conflict_type": "excel_alquiler_multiple",
                "serial_busqueda": candidate.get("numero_serie") or "",
                "serial_encontrado": candidate_serial_norm,
                "os_access": "",
                "detalle": f"Se encontraron {len(rental_matches)} coincidencias de alquiler en Excel.",
                "payload_json": "",
            }
        )
    if len(loan_matches) > 1:
        conflicts.append(
            {
                "ingreso_id": candidate["ingreso_id"],
                "device_id": candidate["device_id"],
                "conflict_type": "excel_prestamo_multiple",
                "serial_busqueda": candidate.get("numero_serie") or "",
                "serial_encontrado": candidate_serial_norm,
                "os_access": "",
                "detalle": f"Se encontraron {len(loan_matches)} coincidencias de préstamo en Excel.",
                "payload_json": "",
            }
        )
    return {
        "sales": sale_matches,
        "rentals": rental_matches,
        "loans": loan_matches,
        "conflicts": conflicts,
    }


def _history_rows_for_ingreso(ingreso_id: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if _table_exists("change_log", schema="audit"):
        try:
            rows.extend(
                _fetchall_dicts(
                    """
                    SELECT id, ts, user_id, user_role, table_name, record_id, column_name, old_value, new_value
                      FROM audit.change_log
                     WHERE ingreso_id = %s
                     ORDER BY ts DESC, id DESC
                     LIMIT 50
                    """,
                    [ingreso_id],
                )
            )
        except Exception:
            pass
    if _table_exists("audit_log"):
        try:
            rows.extend(
                _fetchall_dicts(
                    """
                    SELECT id, ts, user_id, role AS user_role, method, path, body
                      FROM audit_log
                     WHERE path LIKE %s OR path = %s OR path LIKE %s OR path = %s
                     ORDER BY ts DESC, id DESC
                     LIMIT 50
                    """,
                    [
                        f"/api/ingresos/{ingreso_id}/%",
                        f"/api/ingresos/{ingreso_id}/",
                        f"/api/quotes/{ingreso_id}/%",
                        f"/api/quotes/{ingreso_id}/",
                    ],
                )
            )
        except Exception:
            pass
    return rows


def collect_internal_evidence(candidate: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    ingreso_id = int(candidate["ingreso_id"])
    device_id = int(candidate["device_id"])
    evidences: list[dict[str, Any]] = []
    summary = {
        "sale_event": False,
        "sale_event_date": None,
        "sale_event_factura": "",
        "sale_event_remito": "",
        "baja_historical": False,
        "baja_historical_date": None,
        "alquiler_historical": False,
        "alquiler_historical_date": None,
        "alquiler_historical_payload": {},
        "terminal_event": False,
        "terminal_event_state": "",
        "terminal_event_date": None,
    }

    if _table_exists("device_mg_events"):
        try:
            rows = _fetchall_dicts(
                "SELECT * FROM device_mg_events WHERE device_id=%s ORDER BY id DESC LIMIT 20",
                [device_id],
            )
            rows.sort(key=lambda row: _order_key_from_row(row, "fecha_evento", "created_at"), reverse=True)
            for row in rows:
                accion = str(row.get("accion") or "").strip().lower()
                fecha_evento = _parse_effective_datetime(row.get("fecha_evento") or row.get("created_at"))
                if accion == "venta" and not summary["sale_event"]:
                    summary["sale_event"] = True
                    summary["sale_event_date"] = fecha_evento
                    summary["sale_event_factura"] = str(row.get("factura_numero") or "").strip()
                    summary["sale_event_remito"] = str(row.get("remito_numero") or "").strip()
                evidences.append(
                    {
                        "ingreso_id": ingreso_id,
                        "device_id": device_id,
                        "source": "device_mg_events",
                        "source_priority": 3,
                        "match_type": accion or "evento",
                        "serial_busqueda": candidate.get("numero_serie") or "",
                        "serial_encontrado": candidate.get("numero_serie") or "",
                        "fecha_fuente": _format_scalar(fecha_evento),
                        "factura": str(row.get("factura_numero") or "").strip(),
                        "remito": str(row.get("remito_numero") or "").strip(),
                        "deposito": "",
                        "os_access": "",
                        "confidence_hint": "alta" if accion == "venta" else "media",
                        "detalle": str(row.get("observaciones") or row.get("source") or "Evento MG"),
                    }
                )
        except Exception:
            pass

    if _table_exists("ingreso_historical_corrections"):
        try:
            rows = _fetchall_dicts(
                "SELECT * FROM ingreso_historical_corrections WHERE ingreso_id=%s ORDER BY fecha_efectiva DESC, id DESC LIMIT 20",
                [ingreso_id],
            )
            for row in rows:
                accion = str(row.get("accion") or "").strip().lower()
                fecha_efectiva = _parse_effective_datetime(row.get("fecha_efectiva"))
                if accion == "baja_ingreso" and not summary["baja_historical"]:
                    summary["baja_historical"] = True
                    summary["baja_historical_date"] = fecha_efectiva
                if accion == "alta_alquiler" and not summary["alquiler_historical"]:
                    summary["alquiler_historical"] = True
                    summary["alquiler_historical_date"] = fecha_efectiva
                    payload = row.get("payload")
                    if isinstance(payload, str):
                        try:
                            payload = json.loads(payload)
                        except Exception:
                            payload = {}
                    summary["alquiler_historical_payload"] = payload or {}
                evidences.append(
                    {
                        "ingreso_id": ingreso_id,
                        "device_id": device_id,
                        "source": "ingreso_historical_corrections",
                        "source_priority": 3,
                        "match_type": accion or "correccion",
                        "serial_busqueda": candidate.get("numero_serie") or "",
                        "serial_encontrado": candidate.get("numero_serie") or "",
                        "fecha_fuente": _format_scalar(fecha_efectiva),
                        "factura": "",
                        "remito": "",
                        "deposito": "",
                        "os_access": "",
                        "confidence_hint": "alta" if accion == "baja_ingreso" else "media",
                        "detalle": str(row.get("motivo") or "Corrección histórica"),
                    }
                )
        except Exception:
            pass

    if _table_exists("ingreso_events"):
        ticket_column = "ticket_id" if _table_has_column("ingreso_events", "ticket_id") else "ingreso_id"
        try:
            rows = _fetchall_dicts(
                f"SELECT * FROM ingreso_events WHERE {ticket_column}=%s ORDER BY ts DESC, id DESC LIMIT 30",
                [ingreso_id],
            )
            for row in rows:
                new_state = str(row.get("a_estado") or "").strip().lower()
                timestamp = _parse_effective_datetime(row.get("ts"))
                if new_state in TERMINAL_STATES and not summary["terminal_event"]:
                    summary["terminal_event"] = True
                    summary["terminal_event_state"] = new_state
                    summary["terminal_event_date"] = timestamp
                evidences.append(
                    {
                        "ingreso_id": ingreso_id,
                        "device_id": device_id,
                        "source": "ingreso_events",
                        "source_priority": 3,
                        "match_type": new_state or "evento_estado",
                        "serial_busqueda": candidate.get("numero_serie") or "",
                        "serial_encontrado": candidate.get("numero_serie") or "",
                        "fecha_fuente": _format_scalar(timestamp),
                        "factura": "",
                        "remito": "",
                        "deposito": "",
                        "os_access": "",
                        "confidence_hint": "media",
                        "detalle": str(row.get("comentario") or ""),
                    }
                )
        except Exception:
            pass

    if _table_exists("bejerman_sync_jobs"):
        try:
            rows = _fetchall_dicts(
                """
                SELECT *
                  FROM bejerman_sync_jobs
                 WHERE ingreso_id=%s OR device_id=%s
                 ORDER BY updated_at DESC, id DESC
                 LIMIT 20
                """,
                [ingreso_id, device_id],
            )
            for row in rows:
                evidences.append(
                    {
                        "ingreso_id": ingreso_id,
                        "device_id": device_id,
                        "source": "bejerman_sync_jobs",
                        "source_priority": 3,
                        "match_type": str(row.get("sync_type") or "").strip(),
                        "serial_busqueda": candidate.get("numero_serie") or "",
                        "serial_encontrado": str(row.get("numero_serie") or "").strip(),
                        "fecha_fuente": _format_scalar(
                            _parse_effective_datetime(row.get("updated_at") or row.get("created_at"))
                        ),
                        "factura": "",
                        "remito": "",
                        "deposito": str(row.get("target_deposit") or row.get("source_deposit") or "").strip(),
                        "os_access": "",
                        "confidence_hint": "baja",
                        "detalle": str(row.get("status") or "").strip(),
                    }
                )
        except Exception:
            pass

    for row in _history_rows_for_ingreso(ingreso_id):
        detail = str(row.get("column_name") or row.get("path") or "auditoría")
        if row.get("new_value") not in (None, ""):
            detail = f"{detail}: {row.get('new_value')}"
        evidences.append(
            {
                "ingreso_id": ingreso_id,
                "device_id": device_id,
                "source": "audit",
                "source_priority": 3,
                "match_type": str(row.get("table_name") or row.get("method") or "audit"),
                "serial_busqueda": candidate.get("numero_serie") or "",
                "serial_encontrado": candidate.get("numero_serie") or "",
                "fecha_fuente": _format_scalar(_parse_effective_datetime(row.get("ts"))),
                "factura": "",
                "remito": "",
                "deposito": "",
                "os_access": "",
                "confidence_hint": "baja",
                "detalle": detail[:512],
            }
        )

    return evidences, summary


def _candidate_mg_codes(candidate: dict[str, Any]) -> list[str]:
    codes: list[str] = []
    for value in (candidate.get("numero_interno"), candidate.get("n_de_control")):
        code = _norm_stockcode(value)
        if code and code not in codes:
            codes.append(code)
    return codes


def _index_access_fallback_rows(rows: Sequence[dict[str, Any]]) -> dict[str, dict[Any, list[dict[str, Any]]]]:
    by_remito: dict[int, list[dict[str, Any]]] = defaultdict(list)
    by_mg: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        for remito_field in ("RemitoIngreso", "Remito"):
            remito = _parse_int(row.get(remito_field))
            if remito is not None:
                by_remito[remito].append(row)
        for mg_field in ("NumeroSerie", "NdeControl"):
            mg_code = _norm_stockcode(row.get(mg_field))
            if mg_code:
                by_mg[mg_code].append(row)
    return {"by_remito": by_remito, "by_mg": by_mg}


def _select_access_row(candidate: dict[str, Any], exact_row: Optional[dict[str, Any]], fallback_index: dict[str, dict[Any, list[dict[str, Any]]]]) -> Optional[dict[str, Any]]:
    if exact_row:
        return exact_row
    remito_ingreso = _parse_int(candidate.get("remito_ingreso"))
    if remito_ingreso is None:
        return None
    candidate_mg_codes = _candidate_mg_codes(candidate)
    candidate_brand = _norm_upper_ascii(candidate.get("marca") or "")
    candidate_model = _norm_upper_ascii(candidate.get("modelo") or "")
    best_row: Optional[dict[str, Any]] = None
    best_score = -1
    for row in fallback_index.get("by_remito", {}).get(remito_ingreso, []):
        score = 100
        row_numero_serie_mg = _norm_stockcode(row.get("NumeroSerie"))
        row_ndecontrol_mg = _norm_stockcode(row.get("NdeControl"))
        if row_numero_serie_mg in candidate_mg_codes:
            score += 50
        if row_ndecontrol_mg in candidate_mg_codes:
            score += 40
        if candidate_brand and _norm_upper_ascii(row.get("Marca") or "") == candidate_brand:
            score += 15
        if candidate_model and _norm_upper_ascii(row.get("Modelo") or "") == candidate_model:
            score += 10
        if score > best_score:
            best_score = score
            best_row = row
    if best_row and best_score >= 140:
        shifted_row = dict(best_row)
        shifted_row["__trace_match_strategy"] = "remito_ingreso_shifted"
        shifted_row["__trace_match_detail"] = (
            f"Access coincide por remito de ingreso {remito_ingreso} y MG; la OS en Access figura como {best_row.get('Id')}."
        )
        shifted_row["__trace_serial_encontrado"] = str(best_row.get("NumeroSerie") or best_row.get("NdeControl") or "").strip()
        return shifted_row
    return None


def _access_status_summary(row: Optional[dict[str, Any]]) -> dict[str, Any]:
    if not row:
        return {
            "venta": False,
            "entregado": False,
            "alquilado": False,
            "baja_markers": [],
            "recibido": "",
            "comentarios": "",
            "alquiler_destino": "",
            "fecha_referencia": None,
        }
    recibido = str(row.get("Recibido") or "").strip()
    comentarios = str(row.get("Comentarios") or "").strip()
    baja_markers: list[str] = []
    for label, value in (("Recibido", recibido), ("Comentarios", comentarios)):
        if value and "BAJA" in _norm_upper_ascii(value):
            baja_markers.append(f"{label}={value}")
    alquiler_destino = str(row.get("CargoAlquiler") or row.get("RecibeAlquiler") or "").strip()
    fecha_referencia = _parse_effective_datetime(row.get("FechaEntrega") or row.get("FechaIngreso"))
    return {
        "venta": _parse_bool(row.get("Venta")),
        "entregado": _parse_bool(row.get("Entregado")),
        "alquilado": _parse_bool(row.get("Alquilado")),
        "baja_markers": baja_markers,
        "recibido": recibido,
        "comentarios": comentarios,
        "alquiler_destino": alquiler_destino,
        "fecha_referencia": fecha_referencia,
    }


def _access_match_context(candidate: dict[str, Any], row: dict[str, Any]) -> dict[str, Any]:
    trace_strategy = str(row.get("__trace_match_strategy") or "").strip()
    if trace_strategy == "remito_ingreso_shifted":
        return {
            "match_type": "remito_ingreso_mg_shifted",
            "match_basis": "remito_ingreso",
            "serial_encontrado": str(row.get("__trace_serial_encontrado") or row.get("NumeroSerie") or row.get("NdeControl") or "").strip(),
            "detail": str(row.get("__trace_match_detail") or "Access coincide por remito de ingreso y MG con OS desplazada.").strip(),
            "conflict": None,
        }
    serial_access = str(row.get("NumeroSerie") or "").strip()
    ndecontrol_access = str(row.get("NdeControl") or "").strip()
    serial_norm = _norm_ns(serial_access)
    candidate_serial_norm = _norm_ns(candidate.get("numero_serie"))
    serial_stockcode = _norm_stockcode(serial_access)
    ndecontrol_stockcode = _norm_stockcode(ndecontrol_access)
    candidate_mg_codes = _candidate_mg_codes(candidate)
    if serial_norm and candidate_serial_norm and serial_norm == candidate_serial_norm:
        return {
            "match_type": "os_exact_serial",
            "match_basis": "numero_serie",
            "serial_encontrado": serial_access,
            "detail": "Access coincide por OS y por número de serie.",
            "conflict": None,
        }
    if serial_stockcode and serial_stockcode in candidate_mg_codes:
        return {
            "match_type": "os_exact_mg_en_numero_serie",
            "match_basis": "numero_interno_mg",
            "serial_encontrado": serial_access,
            "detail": f"Access coincide por OS y usa el MG {serial_stockcode} dentro de NumeroSerie.",
            "conflict": None,
        }
    if ndecontrol_stockcode and ndecontrol_stockcode in candidate_mg_codes:
        return {
            "match_type": "os_exact_mg_en_ndecontrol",
            "match_basis": "numero_interno_mg",
            "serial_encontrado": ndecontrol_access,
            "detail": f"Access coincide por OS y usa el MG {ndecontrol_stockcode} dentro de NdeControl.",
            "conflict": None,
        }
    if not serial_access and not ndecontrol_access:
        return {
            "match_type": "os_exact_sin_serie_access",
            "match_basis": "os",
            "serial_encontrado": "",
            "detail": "Access coincide por OS pero no informa número de serie/MG en la fila.",
            "conflict": None,
        }
    return {
        "match_type": "os_exact_serial_mismatch",
        "match_basis": "os",
        "serial_encontrado": serial_access or ndecontrol_access,
        "detail": "Access coincide por OS, pero no coincide ni por número de serie ni por MG del SR.",
        "conflict": {
            "ingreso_id": candidate["ingreso_id"],
            "device_id": candidate["device_id"],
            "conflict_type": "access_serial_mismatch",
            "serial_busqueda": candidate.get("numero_serie") or "",
            "serial_encontrado": serial_access or ndecontrol_access,
            "os_access": row.get("Id") or "",
            "detalle": "Access coincide por OS, pero no coincide ni por número de serie ni por MG del SR.",
            "payload_json": "",
        },
    }


def _build_access_evidence(candidate: dict[str, Any], row: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    evidences: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    row_id = _parse_int(row.get("Id"))
    trace_strategy = str(row.get("__trace_match_strategy") or "").strip()
    if row_id != int(candidate["ingreso_id"]) and trace_strategy != "remito_ingreso_shifted":
        conflicts.append(
            {
                "ingreso_id": candidate["ingreso_id"],
                "device_id": candidate["device_id"],
                "conflict_type": "access_os_mismatch",
                "serial_busqueda": candidate.get("numero_serie") or "",
                "serial_encontrado": str(row.get("NumeroSerie") or "").strip(),
                "os_access": row.get("Id") or "",
                "detalle": "La fila devuelta por Access no coincide con la OS solicitada.",
                "payload_json": "",
            }
        )
        return evidences, conflicts
    match = _access_match_context(candidate, row)
    if match.get("conflict"):
        conflicts.append(match["conflict"])
    access_status = _access_status_summary(row)
    detail_parts = []
    detail_parts.append(match["detail"])
    if access_status["venta"]:
        detail_parts.append("Venta=Sí")
    if access_status["entregado"]:
        detail_parts.append("Entregado=Sí")
    if access_status["alquilado"]:
        detail_parts.append("Alquilado=Sí")
    estado = str(row.get("Estado") or "").strip()
    if estado:
        detail_parts.append(f"Estado={estado}")
    if access_status["recibido"]:
        detail_parts.append(f"Recibido={access_status['recibido']}")
    if access_status["comentarios"]:
        detail_parts.append(f"Comentarios={access_status['comentarios']}")
    if access_status["alquiler_destino"]:
        detail_parts.append(f"DestinoAlquiler={access_status['alquiler_destino']}")
    evidences.append(
        {
            "ingreso_id": candidate["ingreso_id"],
            "device_id": candidate["device_id"],
            "source": "access_servicio",
            "source_priority": 4,
            "match_type": match["match_type"],
            "serial_busqueda": candidate.get("numero_serie") or "",
            "serial_encontrado": match["serial_encontrado"],
            "fecha_fuente": str(row.get("FechaEntrega") or row.get("FechaIngreso") or "").strip(),
            "factura": str(row.get("Factura") or "").strip(),
            "remito": str(row.get("Remito") or "").strip(),
            "deposito": "",
            "os_access": row.get("Id") or "",
            "confidence_hint": "baja",
            "classification": "access_os_match",
            "conclusion_operativa": "baja_access" if access_status["baja_markers"] else "access_complementario",
            "detalle": " | ".join(detail_parts) or "Registro Access",
        }
    )
    return evidences, conflicts


def _safe_record_payload(record: WorkbookRecord) -> dict[str, Any]:
    return {
        "sheet_key": record.sheet_key,
        "source_name": record.source_name,
        "row_number": record.row_number,
        "codigo_gen": record.codigo_gen,
        "fecha": _format_scalar(record.fecha),
        "comprobante": record.comprobante,
        "equipo": record.equipo,
        "serial_raw": record.serial_raw,
        "serial_norm": record.serial_norm,
        "precio_neto": record.precio_neto,
        "iva": record.iva,
        "remito": record.remito,
    }


def _summarize_bejerman_evidences(evidences: Sequence[dict[str, Any]]) -> dict[str, Any]:
    deposits = sorted({str(item.get("deposito") or "").strip() for item in evidences if str(item.get("deposito") or "").strip()})
    conclusions = {str(item.get("conclusion_operativa") or "").strip() for item in evidences}
    if not evidences:
        return {"kind": "", "reason": "", "deposits": deposits}
    if conclusions == {"sin_indicio_salida"}:
        joined = ", ".join(deposits) if deposits else "-"
        return {
            "kind": "stock_servicio",
            "reason": f"Bejerman ubica la partida en {joined}; no hay indicio de venta ni de alquiler.",
            "deposits": deposits,
        }
    if "indicio_alquiler" in conclusions:
        joined = ", ".join(deposits) if deposits else "STL"
        return {
            "kind": "indicio_alquiler",
            "reason": f"Bejerman ubica la partida en {joined}; es un indicio de stock de alquiler MG, no de venta.",
            "deposits": deposits,
        }
    if "indicio_entrega_cliente" in conclusions:
        joined = ", ".join(deposits) if deposits else "STC"
        return {
            "kind": "indicio_entrega_cliente",
            "reason": f"Bejerman ubica la partida en {joined}; sugiere entrega/salida a cliente, pero no define venta o devolución.",
            "deposits": deposits,
        }
    if "indicio_salida_final" in conclusions:
        joined = ", ".join(deposits) if deposits else "SALIDA"
        return {
            "kind": "indicio_salida_final",
            "reason": f"Bejerman registra la partida en {joined}; hay una salida, pero no alcanza para distinguir venta, alquiler o devolución.",
            "deposits": deposits,
        }
    joined = ", ".join(deposits) if deposits else "-"
    return {
        "kind": "indeterminado",
        "reason": f"Bejerman encontró la partida en {joined}; requiere interpretación manual.",
        "deposits": deposits,
    }


def _should_hide_candidate_from_outputs(candidate: dict[str, Any], proposal: dict[str, Any]) -> bool:
    reason = str(proposal.get("reason") or "").strip()
    location = _norm_upper_ascii(candidate.get("ubicacion_nombre") or "")
    return reason == "Sin evidencia concluyente." and "ESTANTERIA DE ALQUILER" in location


def _row_ingreso_device_key(row: dict[str, Any]) -> tuple[Optional[int], Optional[int]]:
    return _parse_int(row.get("ingreso_id")), _parse_int(row.get("device_id"))


def build_proposal(
    candidate: dict[str, Any],
    workbook_matches: dict[str, Any],
    internal_summary: dict[str, Any],
    access_row: Optional[dict[str, Any]],
    bejerman_evidences: list[dict[str, Any]],
    duplicate_serials: dict[str, list[int]],
    sale_customer_id: Optional[int],
) -> dict[str, Any]:
    candidate_serial_norm = _norm_ns(candidate.get("numero_serie"))
    duplicate_list = duplicate_serials.get(candidate_serial_norm, [])
    has_duplicate_conflict = len(duplicate_list) > 1
    sale_matches = workbook_matches["sales"]
    rental_matches = workbook_matches["rentals"]
    loan_matches = workbook_matches["loans"]
    access_status = _access_status_summary(access_row)
    bejerman_summary = _summarize_bejerman_evidences(bejerman_evidences)

    proposal = {
        "ingreso_id": candidate["ingreso_id"],
        "device_id": candidate["device_id"],
        "approved": 0,
        "action": "sin_cambios",
        "confidence": "baja",
        "reason": "Sin evidencia concluyente.",
        "source_summary": "",
        "fecha_efectiva": "",
        "fecha_venta": "",
        "venta_customer_id": sale_customer_id or "",
        "venta_customer_nombre": DEFAULT_SALE_CUSTOMER_NAME if sale_matches else "",
        "venta_numero_alternativo": candidate.get("numero_interno") or "",
        "factura_numero": "",
        "remito_numero": "",
        "alquiler_a": "",
        "alquiler_fecha": "",
        "alquiler_remito": "",
        "notificar": 0,
        "payload_json": "",
    }

    if has_duplicate_conflict:
        proposal["action"] = "revision_manual"
        proposal["confidence"] = "baja"
        proposal["reason"] = f"Número de serie duplicado en {len(duplicate_list)} ingresos abiertos."
        proposal["source_summary"] = "conflicto_serial_duplicado"
        proposal["payload_json"] = _json_dumps(
            {
                "duplicate_ingresos": duplicate_list,
                "serial_norm": candidate_serial_norm,
            }
        )
        return proposal

    if loan_matches:
        record = loan_matches[0]["record"]
        corroborated_date = internal_summary["alquiler_historical_date"] or access_status["fecha_referencia"]
        payload = {
            "action": "alquilado",
            "accion": "alta_alquiler",
            "motivo": "Regularización histórica por préstamo a NOVAMED SA; registrar como alquiler.",
            "fecha_efectiva": _format_scalar(corroborated_date),
            "alquiler_a": DEFAULT_SALE_CUSTOMER_NAME,
            "alquiler_fecha": _format_scalar(corroborated_date.date() if hasattr(corroborated_date, "date") else corroborated_date),
            "alquiler_remito": "",
            "notificar": False,
            "observacion": "Coincidencia en hoja PRÉSTAMO; tratar como alquiler a NOVAMED SA.",
            "prestamo": _safe_record_payload(record),
        }
        proposal["action"] = "alquilado"
        proposal["confidence"] = "media"
        proposal["reason"] = (
            "Coincidencia en la hoja PRÉSTAMO; tratar como alquiler a NOVAMED SA."
            if corroborated_date
            else "Coincidencia en la hoja PRÉSTAMO; tratar como alquiler a NOVAMED SA, pero falta fecha efectiva para aplicar."
        )
        proposal["source_summary"] = f"excel_{record.sheet_key}"
        proposal["fecha_efectiva"] = _format_scalar(corroborated_date)
        proposal["alquiler_a"] = DEFAULT_SALE_CUSTOMER_NAME
        proposal["alquiler_fecha"] = payload["alquiler_fecha"]
        proposal["payload_json"] = _json_dumps(payload)
        return proposal

    if sale_matches:
        match = sale_matches[0]
        record: WorkbookRecord = match["record"]
        serial_info = match["serial_info"]
        confidence = "alta" if serial_info["match_type"] == "serial_exact" else "media"
        payload = {
            "action": "vendido_entregado",
            "motivo": DEFAULT_SALE_REASON,
            "fecha_venta": _format_scalar(record.fecha),
            "fecha_efectiva": _format_scalar(record.fecha),
            "factura_numero": record.comprobante,
            "remito_numero": record.remito,
            "venta_customer_id": sale_customer_id,
            "venta_customer_nombre": DEFAULT_SALE_CUSTOMER_NAME,
            "venta_numero_alternativo": candidate.get("numero_interno") or "",
            "source": "trace_pending_devices",
            "trace": _safe_record_payload(record),
        }
        proposal.update(
            {
                "action": "vendido_entregado",
                "confidence": confidence,
                "reason": f"Coincidencia de venta en {record.sheet_key}.",
                "source_summary": f"excel_{record.sheet_key}",
                "fecha_efectiva": _format_scalar(record.fecha),
                "fecha_venta": _format_scalar(record.fecha),
                "factura_numero": record.comprobante,
                "remito_numero": record.remito,
                "payload_json": _json_dumps(payload),
            }
        )
        return proposal

    if internal_summary["baja_historical"]:
        payload = {
            "action": "baja",
            "accion": "baja_ingreso",
            "motivo": DEFAULT_BAJA_REASON,
            "fecha_efectiva": _format_scalar(internal_summary["baja_historical_date"]),
            "notificar": False,
        }
        proposal.update(
            {
                "action": "baja",
                "confidence": "alta",
                "reason": "Existe corrección histórica de baja en el sistema.",
                "source_summary": "ingreso_historical_corrections",
                "fecha_efectiva": _format_scalar(internal_summary["baja_historical_date"]),
                "payload_json": _json_dumps(payload),
            }
        )
        return proposal

    if internal_summary["sale_event"] and str(candidate.get("mg_estado") or "").strip().lower() == "inactivo_venta":
        payload = {
            "action": "vendido_entregado",
            "motivo": DEFAULT_SALE_REASON,
            "fecha_venta": _format_scalar(internal_summary["sale_event_date"] or candidate.get("mg_venta_fecha")),
            "fecha_efectiva": _format_scalar(internal_summary["sale_event_date"] or candidate.get("mg_venta_fecha")),
            "factura_numero": internal_summary["sale_event_factura"] or candidate.get("mg_venta_factura_numero") or "",
            "remito_numero": internal_summary["sale_event_remito"] or candidate.get("mg_venta_remito_numero") or "",
            "venta_customer_id": candidate.get("mg_venta_customer_id") or sale_customer_id,
            "venta_customer_nombre": candidate.get("mg_venta_customer_nombre") or DEFAULT_SALE_CUSTOMER_NAME,
            "venta_numero_alternativo": candidate.get("mg_venta_numero_alternativo") or candidate.get("numero_interno") or "",
            "source": "trace_pending_devices",
        }
        proposal.update(
            {
                "action": "vendido_entregado",
                "confidence": "alta",
                "reason": "Equipo marcado como vendido en trazas internas consistentes.",
                "source_summary": "device_mg_events",
                "fecha_efectiva": payload["fecha_efectiva"],
                "fecha_venta": payload["fecha_venta"],
                "factura_numero": payload["factura_numero"],
                "remito_numero": payload["remito_numero"],
                "venta_customer_id": payload["venta_customer_id"] or "",
                "venta_customer_nombre": payload["venta_customer_nombre"] or "",
                "venta_numero_alternativo": payload["venta_numero_alternativo"] or "",
                "payload_json": _json_dumps(payload),
            }
        )
        return proposal

    if rental_matches:
        corroborated_date = None
        alquiler_a = ""
        alquiler_remito = ""
        if internal_summary["alquiler_historical"]:
            corroborated_date = internal_summary["alquiler_historical_date"]
            alquiler_a = str(
                internal_summary["alquiler_historical_payload"].get("alquiler_a")
                or candidate.get("cliente")
                or ""
            ).strip()
            alquiler_remito = str(
                internal_summary["alquiler_historical_payload"].get("alquiler_remito") or ""
            ).strip()
        elif access_status["alquilado"]:
            corroborated_date = access_status["fecha_referencia"]
            alquiler_a = access_status["alquiler_destino"]
        if corroborated_date:
            payload = {
                "action": "alquilado",
                "accion": "alta_alquiler",
                "motivo": DEFAULT_RENT_REASON,
                "fecha_efectiva": _format_scalar(corroborated_date),
                "alquiler_a": alquiler_a,
                "alquiler_fecha": _format_scalar(corroborated_date.date() if hasattr(corroborated_date, "date") else corroborated_date),
                "alquiler_remito": alquiler_remito,
                "notificar": False,
            }
            proposal.update(
                {
                    "action": "alquilado",
                    "confidence": "media",
                    "reason": "Coincidencia de alquiler corroborada con trazas internas o Access.",
                    "source_summary": "excel_alquiler",
                    "fecha_efectiva": payload["fecha_efectiva"],
                    "alquiler_a": alquiler_a,
                    "alquiler_fecha": payload["alquiler_fecha"],
                    "alquiler_remito": alquiler_remito,
                    "payload_json": _json_dumps(payload),
                }
            )
            return proposal
        proposal.update(
            {
                "action": "revision_manual",
                "confidence": "media",
                "reason": "Coincidencia en alquiler NOVAMED sin fecha efectiva confiable para autoaplicar.",
                "source_summary": "excel_alquiler",
                "payload_json": _json_dumps({"trace": _safe_record_payload(rental_matches[0]["record"])}),
            }
        )
        return proposal

    if access_row:
        if access_status["baja_markers"]:
            payload = {
                "action": "baja",
                "accion": "baja_ingreso",
                "motivo": "Access indica BAJA en la hoja técnica histórica.",
                "fecha_efectiva": _format_scalar(access_status["fecha_referencia"]),
                "notificar": False,
                "access": access_row,
            }
            proposal.update(
                {
                    "action": "baja",
                    "confidence": "baja",
                    "reason": f"Access indica BAJA ({'; '.join(access_status['baja_markers'])}); sugiere marcar baja en SR tras revisión.",
                    "source_summary": "access_servicio",
                    "fecha_efectiva": _format_scalar(access_status["fecha_referencia"]),
                    "payload_json": _json_dumps(payload),
                }
            )
            return proposal
        if access_status["venta"] or access_status["entregado"] or access_status["alquilado"]:
            if access_status["venta"]:
                reason = "Access marca Venta=Sí; revisar venta en SR."
            elif access_status["alquilado"]:
                reason = "Access marca Alquilado=Sí; revisar alquiler en SR."
            else:
                reason = "Access marca Entregado=Sí; revisar salida/entrega en SR."
            proposal.update(
                {
                    "action": "revision_manual",
                    "confidence": "baja",
                    "reason": reason,
                    "source_summary": "access_servicio",
                    "payload_json": _json_dumps({"access": access_row}),
                }
            )
            return proposal

    if bejerman_evidences:
        action = "revision_manual"
        if bejerman_summary["kind"] == "stock_servicio":
            action = "sin_cambios"
        proposal.update(
            {
                "action": action,
                "confidence": "baja",
                "reason": bejerman_summary["reason"] or "Solo hay evidencia complementaria en Bejerman.",
                "source_summary": "bejerman",
                "payload_json": _json_dumps({"bejerman": bejerman_evidences, "summary": bejerman_summary}),
            }
        )
        return proposal

    return proposal


def trace_pending_devices(
    *,
    cutoff_date: str = DEFAULT_CUTOFF_DATE,
    xlsx_novamed: str = DEFAULT_XLSX_NOVAMED,
    access_db: str = DEFAULT_ACCESS_DB,
    ids: Optional[Sequence[int]] = None,
) -> dict[str, Any]:
    workbook_index = load_novamed_workbook(xlsx_novamed)
    candidates = fetch_candidates(cutoff_date, ids=ids)
    duplicate_serials = build_duplicate_serial_map(candidates)
    access_rows, access_error = load_access_rows(access_db, [int(row["ingreso_id"]) for row in candidates])
    fallback_remitos = sorted(
        {
            int(remito)
            for remito in (_parse_int(row.get("remito_ingreso")) for row in candidates)
            if remito is not None
        }
    )
    fallback_mg_codes = sorted({code for row in candidates for code in _candidate_mg_codes(row)})
    access_fallback_rows, access_fallback_error = load_access_fallback_rows(
        access_db,
        remitos=fallback_remitos,
        mg_codes=fallback_mg_codes,
    )
    access_fallback_index = _index_access_fallback_rows(access_fallback_rows)
    sale_customer_id = _resolve_sale_customer_id(DEFAULT_SALE_CUSTOMER_NAME)
    bejerman_lookup = BejermanLookup()

    outputs = {
        "candidatos": [],
        "evidencias": [],
        "propuestas": [],
        "conflictos": [],
        "prestamo_revision": [],
        "metadata": {
            "cutoff_date": cutoff_date,
            "xlsx_novamed": xlsx_novamed,
            "access_db": access_db,
            "candidate_count_raw": len(candidates),
            "candidate_count": len(candidates),
            "access_error": access_error or "",
            "access_fallback_error": access_fallback_error or "",
            "sale_customer_id": sale_customer_id or "",
            "hidden_estanteria_alquiler_without_evidence": 0,
        },
    }

    for candidate in candidates:
        outputs["candidatos"].append(
            {
                "ingreso_id": candidate["ingreso_id"],
                "device_id": candidate["device_id"],
                "customer_id": candidate.get("customer_id") or "",
                "cliente": candidate.get("cliente") or "",
                "marca": candidate.get("marca") or "",
                "modelo": candidate.get("modelo") or "",
                "tipo_equipo": candidate.get("tipo_equipo") or "",
                "variante": candidate.get("variante") or "",
                "numero_serie": candidate.get("numero_serie") or "",
                "numero_interno": candidate.get("numero_interno") or "",
                "n_de_control": candidate.get("n_de_control") or "",
                "remito_ingreso": candidate.get("remito_ingreso") or "",
                "estado_actual": candidate.get("estado_actual") or "",
                "fecha_base": _format_scalar(candidate.get("fecha_base")),
                "mg_estado": candidate.get("mg_estado") or "",
                "mg_venta_fecha": _format_scalar(candidate.get("mg_venta_fecha")),
                "mg_venta_factura_numero": candidate.get("mg_venta_factura_numero") or "",
                "mg_venta_remito_numero": candidate.get("mg_venta_remito_numero") or "",
                "mg_venta_customer_id": candidate.get("mg_venta_customer_id") or "",
                "mg_venta_customer_nombre": candidate.get("mg_venta_customer_nombre") or "",
                "ubicacion_nombre": candidate.get("ubicacion_nombre") or "",
            }
        )

        workbook_matches = collect_workbook_matches(candidate, workbook_index)
        for match in workbook_matches["sales"]:
            confidence = "alta" if match["serial_info"]["match_type"] == "serial_exact" else "media"
            outputs["evidencias"].append(
                _build_workbook_evidence(candidate, match["serial_info"], match["record"], confidence)
            )
        for match in workbook_matches["rentals"]:
            outputs["evidencias"].append(
                _build_workbook_evidence(candidate, match["serial_info"], match["record"], "media")
            )
        for match in workbook_matches["loans"]:
            outputs["evidencias"].append(
                _build_workbook_evidence(candidate, match["serial_info"], match["record"], "media")
            )
        outputs["conflictos"].extend(workbook_matches["conflicts"])

        internal_evidences, internal_summary = collect_internal_evidence(candidate)
        outputs["evidencias"].extend(internal_evidences)

        bejerman_evidences, bejerman_conflicts = bejerman_lookup.lookup_candidate(candidate)
        outputs["evidencias"].extend(bejerman_evidences)
        outputs["conflictos"].extend(bejerman_conflicts)

        access_row = _select_access_row(
            candidate,
            access_rows.get(int(candidate["ingreso_id"])),
            access_fallback_index,
        )
        if access_row:
            access_evidences, access_conflicts = _build_access_evidence(candidate, access_row)
            outputs["evidencias"].extend(access_evidences)
            outputs["conflictos"].extend(access_conflicts)
        elif access_error:
            outputs["conflictos"].append(
                {
                    "ingreso_id": candidate["ingreso_id"],
                    "device_id": candidate["device_id"],
                    "conflict_type": "access_unavailable",
                    "serial_busqueda": candidate.get("numero_serie") or "",
                    "serial_encontrado": "",
                    "os_access": "",
                    "detalle": access_error,
                    "payload_json": "",
                }
            )
        elif access_fallback_error:
            outputs["conflictos"].append(
                {
                    "ingreso_id": candidate["ingreso_id"],
                    "device_id": candidate["device_id"],
                    "conflict_type": "access_fallback_unavailable",
                    "serial_busqueda": candidate.get("numero_serie") or "",
                    "serial_encontrado": "",
                    "os_access": "",
                    "detalle": access_fallback_error,
                    "payload_json": "",
                }
            )

        proposal = build_proposal(
            candidate=candidate,
            workbook_matches=workbook_matches,
            internal_summary=internal_summary,
            access_row=access_row,
            bejerman_evidences=bejerman_evidences,
            duplicate_serials=duplicate_serials,
            sale_customer_id=sale_customer_id,
        )
        candidate_serial = str(candidate.get("numero_serie") or "").strip()
        if not candidate_serial and not access_row:
            outputs["conflictos"].append(
                {
                    "ingreso_id": candidate["ingreso_id"],
                    "device_id": candidate["device_id"],
                    "conflict_type": "numero_serie_ausente",
                    "serial_busqueda": "",
                    "serial_encontrado": "",
                    "os_access": "",
                    "detalle": "El ingreso no tiene número de serie y no puede trazarse por partida.",
                    "payload_json": "",
                }
            )
        outputs["propuestas"].append(proposal)
        for match in workbook_matches["loans"]:
            outputs["prestamo_revision"].append(
                {
                    "ingreso_id": candidate["ingreso_id"],
                    "device_id": candidate["device_id"],
                    "source_sheet": match["record"].sheet_key,
                    "serial_busqueda": match["serial_info"]["serial_busqueda"],
                    "serial_encontrado": match["record"].serial_raw,
                    "action": proposal.get("action") or "",
                    "reason": proposal.get("reason") or "",
                    "fecha_efectiva": proposal.get("fecha_efectiva") or "",
                    "alquiler_a": proposal.get("alquiler_a") or "",
                    "detalle": (
                        f"Coincidencia en hoja PRÉSTAMO, fila {match['record'].row_number}. "
                        "Se trata como alquiler a NOVAMED SA."
                    ),
                    "payload_json": _json_dumps(_safe_record_payload(match["record"])),
                }
            )

    hidden_keys: set[tuple[Optional[int], Optional[int]]] = set()
    for candidate_row in outputs["candidatos"]:
        key = _row_ingreso_device_key(candidate_row)
        proposal = next(
            (item for item in outputs["propuestas"] if _row_ingreso_device_key(item) == key),
            None,
        )
        if proposal and _should_hide_candidate_from_outputs(candidate_row, proposal):
            hidden_keys.add(key)

    if hidden_keys:
        outputs["candidatos"] = [row for row in outputs["candidatos"] if _row_ingreso_device_key(row) not in hidden_keys]
        outputs["propuestas"] = [row for row in outputs["propuestas"] if _row_ingreso_device_key(row) not in hidden_keys]
        outputs["evidencias"] = [row for row in outputs["evidencias"] if _row_ingreso_device_key(row) not in hidden_keys]
        outputs["conflictos"] = [row for row in outputs["conflictos"] if _row_ingreso_device_key(row) not in hidden_keys]
        outputs["prestamo_revision"] = [
            row for row in outputs["prestamo_revision"] if _row_ingreso_device_key(row) not in hidden_keys
        ]
        outputs["metadata"]["hidden_estanteria_alquiler_without_evidence"] = len(hidden_keys)
        outputs["metadata"]["candidate_count"] = len(outputs["candidatos"])

    return outputs


def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key in seen:
                continue
            fieldnames.append(key)
            seen.add(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _format_scalar(row.get(key)) for key in fieldnames})


def _write_xlsx(path: Path, datasets: dict[str, Sequence[dict[str, Any]]]) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in datasets.items():
        ws = workbook.create_sheet(title=sheet_name[:31] or "sheet")
        fieldnames: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row.keys():
                if key in seen:
                    continue
                seen.add(key)
                fieldnames.append(key)
        if not fieldnames:
            fieldnames = ["empty"]
        ws.append(fieldnames)
        for row in rows:
            ws.append([_format_scalar(row.get(key)) for key in fieldnames])
    workbook.save(path)


def write_trace_outputs(out_dir: str | Path, outputs: dict[str, Any]) -> Path:
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    datasets = {
        "candidatos": outputs["candidatos"],
        "evidencias": outputs["evidencias"],
        "propuestas": outputs["propuestas"],
        "conflictos": outputs["conflictos"],
        "prestamo_revision": outputs["prestamo_revision"],
    }
    for name, rows in datasets.items():
        _write_csv(out_path / f"{name}.csv", rows)
    _write_xlsx(out_path / "trace_pending_devices.xlsx", datasets)
    (out_path / "summary.json").write_text(_json_dumps(outputs["metadata"]), encoding="utf-8")
    return out_path


def _load_rows_from_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _load_rows_from_xlsx(path: Path, sheet_name: str = "propuestas") -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return []
        header_list = [str(item or "").strip() for item in headers]
        rows: list[dict[str, Any]] = []
        for row in iterator:
            record = {}
            for idx, key in enumerate(header_list):
                record[key] = row[idx] if idx < len(row) else None
            rows.append(record)
        return rows
    finally:
        workbook.close()


def load_proposal_rows(path: str, sheet_name: str = "propuestas") -> list[dict[str, Any]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _load_rows_from_csv(file_path)
    if suffix in (".xlsx", ".xlsm"):
        return _load_rows_from_xlsx(file_path, sheet_name=sheet_name)
    raise ValueError("Formato no soportado para propuestas. Use .csv o .xlsx.")


def _merge_payload_from_row(row: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    raw_payload = row.get("payload_json")
    if raw_payload not in (None, "", {}):
        try:
            if isinstance(raw_payload, dict):
                payload = dict(raw_payload)
            else:
                payload = json.loads(str(raw_payload))
        except Exception as exc:
            raise ValueError(f"payload_json inválido: {exc}") from exc
    overrides = {
        "fecha_efectiva": row.get("fecha_efectiva"),
        "fecha_venta": row.get("fecha_venta"),
        "venta_customer_id": row.get("venta_customer_id"),
        "venta_customer_nombre": row.get("venta_customer_nombre"),
        "venta_numero_alternativo": row.get("venta_numero_alternativo"),
        "factura_numero": row.get("factura_numero"),
        "remito_numero": row.get("remito_numero"),
        "alquiler_a": row.get("alquiler_a"),
        "alquiler_fecha": row.get("alquiler_fecha"),
        "alquiler_remito": row.get("alquiler_remito"),
    }
    for key, value in overrides.items():
        if value not in (None, ""):
            payload[key] = value
    if row.get("notificar") not in (None, ""):
        payload["notificar"] = _parse_bool(row.get("notificar"))
    return payload


def _current_state_snapshot(ingreso_id: int, device_id: int) -> dict[str, Any]:
    has_mg_state = _table_has_column("devices", "mg_estado")
    has_n_de_control = _table_has_column("devices", "n_de_control")
    sql = f"""
        SELECT
          i.id AS ingreso_id,
          COALESCE(i.estado::text,'') AS ingreso_estado,
          i.fecha_entrega,
          COALESCE(i.factura_numero,'') AS factura_numero,
          COALESCE(i.remito_salida,'') AS remito_salida,
          COALESCE(i.alquilado, FALSE) AS ingreso_alquilado,
          COALESCE(i.alquiler_a,'') AS ingreso_alquiler_a,
          COALESCE(i.alquiler_remito,'') AS ingreso_alquiler_remito,
          i.alquiler_fecha,
          d.id AS device_id,
          COALESCE(d.numero_interno,'') AS numero_interno,
          COALESCE(d.numero_serie,'') AS numero_serie,
          {'COALESCE(d.n_de_control, \'\') AS n_de_control,' if has_n_de_control else '\'\' AS n_de_control,'}
          {'COALESCE(d.mg_estado, \'activo\') AS mg_estado,' if has_mg_state else '\'activo\' AS mg_estado,'}
          COALESCE(d.alquilado, FALSE) AS device_alquilado,
          COALESCE(d.alquiler_a,'') AS device_alquiler_a
          FROM ingresos i
          JOIN devices d ON d.id = i.device_id
         WHERE i.id = %s
           AND d.id = %s
         LIMIT 1
    """
    row = _fetchone_dict(sql, [ingreso_id, device_id])
    if not row:
        raise ValueError(f"No se encontró el ingreso {ingreso_id} / device {device_id}.")
    return row


def _resolve_sale_customer(payload: dict[str, Any]) -> Optional[int]:
    customer_id = _parse_int(payload.get("venta_customer_id"))
    if customer_id:
        return customer_id
    customer_name = str(payload.get("venta_customer_nombre") or DEFAULT_SALE_CUSTOMER_NAME).strip()
    return _resolve_sale_customer_id(customer_name)


def _insert_ingreso_event_safe(
    *,
    ingreso_id: int,
    before_state: str,
    after_state: str,
    actor_user_id: Optional[int],
    timestamp: Any,
    comentario: str,
) -> None:
    try:
        with transaction.atomic():
            _exec(
                """
                INSERT INTO ingreso_events (ticket_id, de_estado, a_estado, usuario_id, ts, comentario)
                VALUES (%s, NULLIF(%s,'')::ticket_state, %s::ticket_state, %s, %s, %s)
                """,
                [ingreso_id, before_state, after_state, actor_user_id, timestamp, comentario],
            )
    except Exception:
        try:
            with transaction.atomic():
                _exec(
                    """
                    INSERT INTO ingreso_events (ticket_id, a_estado, usuario_id, ts, comentario)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [ingreso_id, after_state, actor_user_id, timestamp, comentario],
                )
        except Exception:
            pass


def apply_sale_correction(
    *,
    ingreso_id: int,
    device_id: int,
    payload: dict[str, Any],
    actor_user_id: Optional[int],
) -> dict[str, Any]:
    snapshot = _current_state_snapshot(ingreso_id, device_id)
    fecha_venta = _parse_effective_datetime(payload.get("fecha_venta") or payload.get("fecha_efectiva"))
    if not fecha_venta:
        raise ValueError("fecha_venta inválida para la venta.")
    factura_numero = str(payload.get("factura_numero") or "").strip()
    remito_numero = str(payload.get("remito_numero") or "").strip()
    if not factura_numero and not remito_numero:
        raise ValueError("Debe informar factura o remito para la venta.")
    venta_customer_id = _resolve_sale_customer(payload)
    if not venta_customer_id:
        raise ValueError("No se pudo resolver el comprador de la venta.")
    venta_customer_nombre = str(payload.get("venta_customer_nombre") or DEFAULT_SALE_CUSTOMER_NAME).strip()
    venta_numero_alternativo = str(
        payload.get("venta_numero_alternativo")
        or snapshot.get("numero_interno")
        or snapshot.get("n_de_control")
        or ""
    ).strip()
    observaciones = str(payload.get("motivo") or DEFAULT_SALE_REASON).strip()
    mg_code = _norm_stockcode(snapshot.get("numero_interno") or snapshot.get("n_de_control") or "")
    if not mg_code:
        raise ValueError("El equipo no tiene un número interno MG/NM/NV/CE válido para registrar la venta.")

    dash_id = _dash_location_id()
    before_state = str(snapshot.get("ingreso_estado") or "").strip().lower()
    before_mg_state = str(snapshot.get("mg_estado") or "").strip().lower()

    with transaction.atomic():
        device_sets = [
            "mg_estado = 'inactivo_venta'",
            "mg_inactivo_desde = COALESCE(mg_inactivo_desde, %s)",
            "mg_venta_fecha = %s",
            "mg_venta_factura_numero = NULLIF(%s,'')",
            "mg_venta_remito_numero = NULLIF(%s,'')",
            "mg_venta_observaciones = NULLIF(%s,'')",
            "mg_venta_usuario_id = %s",
            "mg_venta_customer_id = %s",
            "mg_venta_numero_alternativo = NULLIF(%s,'')",
            "alquilado = FALSE",
            "alquiler_a = NULL",
        ]
        device_params: list[Any] = [
            fecha_venta,
            fecha_venta,
            factura_numero,
            remito_numero,
            observaciones,
            actor_user_id,
            venta_customer_id,
            venta_numero_alternativo,
        ]
        if dash_id:
            device_sets.append("ubicacion_id = %s")
            device_params.append(dash_id)
        device_params.append(device_id)
        _exec(
            f"UPDATE devices SET {', '.join(device_sets)} WHERE id = %s",
            device_params,
        )

        _exec(
            """
            INSERT INTO device_mg_events(
              device_id, accion, numero_interno_snapshot, fecha_evento,
              factura_numero, remito_numero, observaciones, usuario_id, ingreso_id,
              venta_customer_id, venta_numero_alternativo, source
            )
            VALUES (%s, 'venta', %s, %s, NULLIF(%s,''), NULLIF(%s,''), NULLIF(%s,''), %s, %s, %s, NULLIF(%s,''), %s)
            """,
            [
                device_id,
                mg_code,
                fecha_venta,
                factura_numero,
                remito_numero,
                observaciones,
                actor_user_id,
                ingreso_id,
                venta_customer_id,
                venta_numero_alternativo,
                "equipos",
            ],
        )

        ingreso_sets = [
            "estado = %s",
            "fecha_entrega = COALESCE(fecha_entrega, %s)",
            "factura_numero = COALESCE(NULLIF(%s,''), factura_numero)",
            "remito_salida = COALESCE(NULLIF(%s,''), remito_salida)",
            "alquilado = FALSE",
            "alquiler_a = NULL",
            "alquiler_remito = NULL",
            "alquiler_fecha = NULL",
        ]
        ingreso_params: list[Any] = [
            "vendido_entregado",
            fecha_venta,
            factura_numero,
            remito_numero,
        ]
        if dash_id:
            ingreso_sets.append("ubicacion_id = %s")
            ingreso_params.append(dash_id)
        ingreso_params.append(ingreso_id)
        _exec(
            f"UPDATE ingresos SET {', '.join(ingreso_sets)} WHERE id = %s",
            ingreso_params,
        )

        _insert_ingreso_event_safe(
            ingreso_id=ingreso_id,
            before_state=before_state,
            after_state="vendido_entregado",
            actor_user_id=actor_user_id,
            timestamp=fecha_venta,
            comentario="Venta registrada y entrega confirmada",
        )

    after_snapshot = _current_state_snapshot(ingreso_id, device_id)
    return {
        "before_ingreso_estado": before_state,
        "after_ingreso_estado": after_snapshot.get("ingreso_estado") or "",
        "before_mg_estado": before_mg_state,
        "after_mg_estado": after_snapshot.get("mg_estado") or "",
        "venta_customer_nombre": venta_customer_nombre,
    }


def apply_historical_correction(
    *,
    ingreso_id: int,
    device_id: int,
    payload: dict[str, Any],
    action: str,
    actor_user_id: Optional[int],
) -> dict[str, Any]:
    snapshot = _current_state_snapshot(ingreso_id, device_id)
    fecha_efectiva = _parse_effective_datetime(payload.get("fecha_efectiva"))
    if not fecha_efectiva:
        raise ValueError("fecha_efectiva inválida.")
    if fecha_efectiva > timezone.now():
        raise ValueError("fecha_efectiva no puede ser futura.")
    motivo = str(payload.get("motivo") or "").strip()
    if not motivo:
        motivo = DEFAULT_RENT_REASON if action == "alquilado" else DEFAULT_BAJA_REASON
    notificar = _parse_bool(payload.get("notificar"), False)
    dash_id = _dash_location_id()
    taller_id = _taller_location_id()
    before_state = str(snapshot.get("ingreso_estado") or "").strip().lower()

    sets: list[str] = []
    params: list[Any] = []
    action_db = ""
    comment = ""
    if action == "alquilado":
        alquiler_a = str(payload.get("alquiler_a") or "").strip()
        alquiler_remito = str(payload.get("alquiler_remito") or "").strip()
        alquiler_fecha = _parse_datetime_or_date(payload.get("alquiler_fecha"))
        alquiler_fecha_date = alquiler_fecha.date() if alquiler_fecha else timezone.localtime(fecha_efectiva).date()
        sets.extend(
            [
                "alquilado=TRUE",
                "estado=%s",
                "alquiler_a=NULLIF(%s,'')",
                "alquiler_remito=NULLIF(%s,'')",
                "alquiler_fecha=%s",
            ]
        )
        params.extend(["alquilado", alquiler_a, alquiler_remito, alquiler_fecha_date])
        if dash_id:
            sets.append("ubicacion_id = COALESCE(%s, ubicacion_id)")
            params.append(dash_id)
        _exec("UPDATE devices SET alquilado=TRUE, alquiler_a=NULLIF(%s,'') WHERE id=%s", [alquiler_a, device_id])
        action_db = "alta_alquiler"
        comment = "Corrección histórica forzada: alta_alquiler"
    elif action == "baja":
        sets.append("estado=%s")
        params.append("baja")
        if dash_id:
            sets.append("ubicacion_id = COALESCE(%s, ubicacion_id)")
            params.append(dash_id)
        action_db = "baja_ingreso"
        comment = "Corrección histórica forzada: baja_ingreso"
    else:
        raise ValueError(f"Acción histórica no soportada: {action}")

    params.append(ingreso_id)
    with transaction.atomic():
        _exec(f"UPDATE ingresos SET {', '.join(sets)} WHERE id=%s", params)
        if action == "baja":
            _exec("UPDATE devices SET alquilado=FALSE, alquiler_a=NULL WHERE id=%s", [device_id])
        if action == "alquilado" and dash_id:
            _exec("UPDATE devices SET ubicacion_id = COALESCE(%s, ubicacion_id) WHERE id=%s", [dash_id, device_id])
        elif action == "baja" and dash_id:
            _exec("UPDATE devices SET ubicacion_id = COALESCE(%s, ubicacion_id) WHERE id=%s", [dash_id, device_id])

        _insert_ingreso_event_safe(
            ingreso_id=ingreso_id,
            before_state=before_state,
            after_state="alquilado" if action == "alquilado" else "baja",
            actor_user_id=actor_user_id,
            timestamp=fecha_efectiva,
            comentario=comment,
        )

        payload_json = _json_dumps(
            {
                **payload,
                "accion": action_db,
                "motivo": motivo,
                "fecha_efectiva": _format_scalar(fecha_efectiva),
                "notificar": notificar,
            }
        )
        if connection.vendor == "postgresql":
            _exec(
                """
                INSERT INTO ingreso_historical_corrections
                  (ingreso_id, accion, fecha_efectiva, motivo, payload, notificar, usuario_id)
                VALUES (%s, %s, %s, %s, %s::jsonb, %s, %s)
                """,
                [ingreso_id, action_db, fecha_efectiva, motivo, payload_json, notificar, actor_user_id],
            )
        else:
            _exec(
                """
                INSERT INTO ingreso_historical_corrections
                  (ingreso_id, accion, fecha_efectiva, motivo, payload, notificar, usuario_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                [ingreso_id, action_db, fecha_efectiva, motivo, payload_json, notificar, actor_user_id],
            )

        if action == "baja" and dash_id:
            _exec("UPDATE devices SET ubicacion_id = COALESCE(%s, ubicacion_id) WHERE id=%s", [dash_id, device_id])

    after_snapshot = _current_state_snapshot(ingreso_id, device_id)
    return {
        "before_ingreso_estado": before_state,
        "after_ingreso_estado": after_snapshot.get("ingreso_estado") or "",
        "before_mg_estado": snapshot.get("mg_estado") or "",
        "after_mg_estado": after_snapshot.get("mg_estado") or "",
    }


def apply_traced_device_corrections(
    *,
    input_path: str,
    actor_email: str = "",
    sheet_name: str = "propuestas",
    dry_run: bool = False,
) -> dict[str, Any]:
    actor_user_id, actor_role = _resolve_actor(actor_email)
    if actor_user_id:
        _set_audit_actor(actor_user_id, actor_role)

    rows = load_proposal_rows(input_path, sheet_name=sheet_name)
    results: list[dict[str, Any]] = []
    for row in rows:
        ingreso_id = _parse_int(row.get("ingreso_id"))
        device_id = _parse_int(row.get("device_id"))
        approved = _parse_bool(row.get("approved"))
        action = str(row.get("action") or "").strip().lower()
        if not ingreso_id or not device_id:
            continue
        if not approved:
            results.append(
                {
                    "ingreso_id": ingreso_id,
                    "device_id": device_id,
                    "action": action or "",
                    "status": "skipped",
                    "detail": "Fila no aprobada.",
                    "before_estado": "",
                    "after_estado": "",
                    "before_mg_estado": "",
                    "after_mg_estado": "",
                    "payload_json": row.get("payload_json") or "",
                }
            )
            continue
        if action not in {"vendido_entregado", "alquilado", "baja"}:
            results.append(
                {
                    "ingreso_id": ingreso_id,
                    "device_id": device_id,
                    "action": action,
                    "status": "skipped",
                    "detail": f"Acción no aplicable en batch: {action}",
                    "before_estado": "",
                    "after_estado": "",
                    "before_mg_estado": "",
                    "after_mg_estado": "",
                    "payload_json": row.get("payload_json") or "",
                }
            )
            continue
        try:
            payload = _merge_payload_from_row(row)
            if dry_run:
                snapshot = _current_state_snapshot(ingreso_id, device_id)
                results.append(
                    {
                        "ingreso_id": ingreso_id,
                        "device_id": device_id,
                        "action": action,
                        "status": "dry_run",
                        "detail": "Fila validada sin aplicar cambios.",
                        "before_estado": snapshot.get("ingreso_estado") or "",
                        "after_estado": action,
                        "before_mg_estado": snapshot.get("mg_estado") or "",
                        "after_mg_estado": "inactivo_venta" if action == "vendido_entregado" else snapshot.get("mg_estado") or "",
                        "payload_json": _json_dumps(payload),
                    }
                )
                continue
            if action == "vendido_entregado":
                applied = apply_sale_correction(
                    ingreso_id=ingreso_id,
                    device_id=device_id,
                    payload=payload,
                    actor_user_id=actor_user_id,
                )
            else:
                applied = apply_historical_correction(
                    ingreso_id=ingreso_id,
                    device_id=device_id,
                    payload=payload,
                    action=action,
                    actor_user_id=actor_user_id,
                )
            results.append(
                {
                    "ingreso_id": ingreso_id,
                    "device_id": device_id,
                    "action": action,
                    "status": "applied",
                    "detail": "",
                    "before_estado": applied.get("before_ingreso_estado") or "",
                    "after_estado": applied.get("after_ingreso_estado") or "",
                    "before_mg_estado": applied.get("before_mg_estado") or "",
                    "after_mg_estado": applied.get("after_mg_estado") or "",
                    "payload_json": _json_dumps(payload),
                }
            )
        except Exception as exc:
            results.append(
                {
                    "ingreso_id": ingreso_id,
                    "device_id": device_id,
                    "action": action,
                    "status": "error",
                    "detail": str(exc),
                    "before_estado": "",
                    "after_estado": "",
                    "before_mg_estado": "",
                    "after_mg_estado": "",
                    "payload_json": row.get("payload_json") or "",
                }
            )

    summary = {
        "input_path": input_path,
        "actor_email": actor_email,
        "dry_run": dry_run,
        "rows_total": len(rows),
        "applied": sum(1 for row in results if row["status"] == "applied"),
        "errors": sum(1 for row in results if row["status"] == "error"),
        "skipped": sum(1 for row in results if row["status"] == "skipped"),
    }
    return {"results": results, "summary": summary}


def write_apply_outputs(out_dir: str | Path, outputs: dict[str, Any]) -> Path:
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    rows = outputs["results"]
    _write_csv(out_path / "applied_results.csv", rows)
    _write_xlsx(out_path / "applied_results.xlsx", {"applied_results": rows})
    (out_path / "summary.json").write_text(_json_dumps(outputs["summary"]), encoding="utf-8")
    return out_path
